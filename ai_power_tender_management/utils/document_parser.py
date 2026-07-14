# Copyright (c) 2026, milind and contributors
# For license information, please see license.txt
"""
Document parsing helpers for the Tender Upload workflow (Phase 1 MVP).

This module only performs *digital* text extraction:
  - PDF  -> pypdf (with PyMuPDF/fitz as a fallback)
  - Excel -> openpyxl

It deliberately does NOT implement OCR. Scanned / image-only PDFs will yield
little or no text; callers should treat that as "OCR Required".

Future AI integration note:
  The real AI pipeline (summarisation, clause detection, structured BOQ
  extraction) will consume the raw text/rows produced here. Keep these
  functions side-effect free so they can be reused by the AI layer later.
"""

import os

import frappe

# Minimum characters of extracted text for a PDF to be considered "readable".
READABLE_TEXT_THRESHOLD = 100


def _resolve_file_path(file_url: str) -> str | None:
	"""Resolve a Frappe file_url (e.g. /private/files/x.pdf) to an absolute path."""
	if not file_url:
		return None

	# Prefer the File doc so both public and private files resolve correctly.
	try:
		file_doc = frappe.get_doc("File", {"file_url": file_url})
		return file_doc.get_full_path()
	except Exception:
		pass

	# Fallback: build the path from the site directory.
	site_path = frappe.get_site_path()
	relative = file_url.lstrip("/")
	candidate = os.path.join(site_path, relative)
	return candidate if os.path.exists(candidate) else None


def extract_text_from_pdf(file_url: str) -> str:
	"""Extract text from a digital PDF. Returns "" if nothing can be read."""
	path = _resolve_file_path(file_url)
	if not path or not os.path.exists(path):
		return ""

	# 1) Try pypdf (pure python, ships with Frappe).
	try:
		from pypdf import PdfReader

		reader = PdfReader(path)
		text = "\n".join((page.extract_text() or "") for page in reader.pages)
		if text.strip():
			return text
	except Exception:
		frappe.log_error(title="Tender Upload: pypdf extract failed", message=frappe.get_traceback())

	# 2) Fallback to PyMuPDF (fitz) if available.
	try:
		import fitz  # PyMuPDF

		doc = fitz.open(path)
		text = "\n".join(page.get_text() for page in doc)
		doc.close()
		if text.strip():
			return text
	except Exception:
		frappe.log_error(title="Tender Upload: fitz extract failed", message=frappe.get_traceback())

	return ""


def is_pdf_text_readable(file_url: str) -> bool:
	"""True when the PDF yields enough digital text to be worth analysing."""
	text = extract_text_from_pdf(file_url)
	return len(text.strip()) > READABLE_TEXT_THRESHOLD


# ---------------------------------------------------------------------------
# OCR (for scanned / image-only PDFs with no text layer)
# ---------------------------------------------------------------------------
import shutil

# Resolve the tesseract binary explicitly — background workers may have a
# minimal PATH that doesn't include Homebrew.
TESSERACT_CMD = shutil.which("tesseract") or "/opt/homebrew/bin/tesseract"


def ocr_available() -> bool:
	"""True when Tesseract + pytesseract are usable."""
	try:
		import pytesseract  # noqa: F401
	except Exception:
		return False
	return bool(shutil.which("tesseract") or os.path.exists(TESSERACT_CMD))


def _configure_tesseract():
	import pytesseract

	pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD
	# Point at the Homebrew tessdata (Arabic pack) if the env isn't already set.
	if "TESSDATA_PREFIX" not in os.environ:
		for candidate in ("/opt/homebrew/share/tessdata", "/usr/local/share/tessdata", "/usr/share/tessdata"):
			if os.path.isdir(candidate):
				os.environ["TESSDATA_PREFIX"] = candidate
				break


def ocr_pdf_pages(file_url: str, langs: str = "ara+eng", dpi: int = 200, page_range=None) -> list[tuple[int, str]]:
	"""
	OCR an image/scanned PDF page-by-page.

	Returns a list of (page_number, text) tuples (page_number is 1-based).
	"""
	path = _resolve_file_path(file_url)
	if not path or not os.path.exists(path):
		return []
	if not ocr_available():
		frappe.log_error(title="Tender OCR: Tesseract not available", message=TESSERACT_CMD)
		return []

	import io

	import fitz
	import pytesseract
	from PIL import Image

	_configure_tesseract()

	out = []
	try:
		doc = fitz.open(path)
	except Exception:
		frappe.log_error(title="Tender OCR: could not open PDF", message=frappe.get_traceback())
		return []

	indexes = range(doc.page_count) if page_range is None else page_range
	for i in indexes:
		try:
			pix = doc[i].get_pixmap(dpi=dpi)
			img = Image.open(io.BytesIO(pix.tobytes("png")))
			text = pytesseract.image_to_string(img, lang=langs)
		except Exception:
			frappe.log_error(title=f"Tender OCR: page {i + 1} failed", message=frappe.get_traceback())
			text = ""
		out.append((i + 1, text))
	doc.close()
	return out


def ocr_pdf_text(file_url: str, langs: str = "ara+eng", dpi: int = 200) -> str:
	"""Full OCR text of a PDF, with [Page N] markers between pages."""
	pages = ocr_pdf_pages(file_url, langs=langs, dpi=dpi)
	return "\n\n".join(f"[Page {n}]\n{t.strip()}" for n, t in pages if t.strip())


# Header keywords used to map spreadsheet columns onto BOQ fields.
_COLUMN_HINTS = {
	"line_type": ["line type", "type", "row type"],
	"item_no": ["item", "item no", "sr", "sr no", "s.no", "no", "sl", "code", "رقم", "رقم البند"],
	"parent_item_no": ["parent item", "parent item no", "parent no", "main item"],
	"description": ["description", "desc", "item description", "particulars", "الوصف", "البيان", "وصف البند", "وصف"],
	"description_en": ["description en", "english description", "description english"],
	"unit": ["unit", "uom", "u.o.m", "الوحدة"],
	"quantity": ["qty", "quantity", "الكمية"],
	"unit_price": ["unit price", "rate", "price", "السعر", "القيمة للوحدة", "قيمة الوحدة", "سعر الوحدة"],
	"specification": ["specification", "spec", "specs", "المواصفات"],
	"source_page": ["source page", "page", "page no"],
	"extraction_confidence": ["confidence", "extraction confidence"],
}


def _match_header(header_cell: str) -> str | None:
	value = (header_cell or "").strip().lower()
	if not value:
		return None
	for field, hints in _COLUMN_HINTS.items():
		if any(value == h or h in value for h in hints):
			return field
	return None


def _detect_header_row(rows, scan_limit: int = 15) -> tuple[int, dict]:
	"""Locate the header row within the first `scan_limit` rows.

	Spreadsheets often start with a merged title or blank rows, so the header
	is not necessarily row 0. Returns (header_index, col_map) for the row that
	maps the most distinct BOQ fields. Requires at least 2 mapped columns to
	qualify as a header; otherwise returns (0, {}).
	"""
	best_idx, best_map = 0, {}
	for idx in range(min(scan_limit, len(rows))):
		col_map = {}
		for cidx, cell in enumerate(rows[idx]):
			field = _match_header(str(cell) if cell is not None else "")
			if field and field not in col_map.values():
				col_map[cidx] = field
		if len(col_map) > len(best_map):
			best_idx, best_map = idx, col_map
	if len(best_map) < 2:
		return 0, {}
	return best_idx, best_map


def excel_to_text_grid(file_url: str, max_rows: int = 500) -> str:
	"""Serialize the first sheet of a workbook to a plain-text grid.

	Each non-empty row becomes a line ``R<n>: c1 | c2 | c3 ...``. NO header
	detection or column mapping is done here — the full grid is handed to the
	LLM, which is far more robust across the many BOQ layouts we receive
	(title rows, merged cells, bilingual headers, subtotals, shifted columns)
	than rule-based column matching. Row indices are preserved so the model
	can cite provenance.
	"""
	path = _resolve_file_path(file_url)
	if not path or not os.path.exists(path):
		return ""
	try:
		import openpyxl

		wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
		ws = wb.active
		lines = []
		for idx, raw in enumerate(ws.iter_rows(values_only=True)):
			if idx >= max_rows:
				break
			cells = ["" if c is None else str(c).strip() for c in raw]
			if any(cells):
				lines.append(f"R{idx}: " + " | ".join(cells))
		wb.close()
		return "\n".join(lines)
	except Exception:
		frappe.log_error(title="Tender Upload: excel_to_text_grid failed", message=frappe.get_traceback())
		return ""


def extract_rows_from_excel(file_url: str) -> list[dict]:
	"""
	Read the first sheet of an Excel workbook and return a list of row dicts
	keyed by BOQ field names.

	Column detection is best-effort based on the header row. If detection
	fails, the raw cell values are returned under generic keys so the caller
	can decide how to handle them.
	"""
	path = _resolve_file_path(file_url)
	if not path or not os.path.exists(path):
		return []

	try:
		import openpyxl

		wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
		ws = wb.active

		rows = list(ws.iter_rows(values_only=True))
		wb.close()
	except Exception:
		frappe.log_error(title="Tender Upload: openpyxl read failed", message=frappe.get_traceback())
		return []

	if not rows:
		return []

	# Detect the header row (may be preceded by title/blank rows) and its
	# column -> field mapping.
	header_idx, col_map = _detect_header_row(rows)

	extracted = []

	if col_map:
		for raw in rows[header_idx + 1:]:
			row = {
				"line_type": None, "item_no": None, "parent_item_no": None,
				"description": None, "description_en": None, "unit": None,
				"quantity": None, "unit_price": None, "specification": None,
				"source_page": None, "extraction_confidence": None,
			}
			has_value = False
			for idx, field in col_map.items():
				if idx < len(raw) and raw[idx] not in (None, ""):
					row[field] = raw[idx]
					has_value = True
			if has_value:
				extracted.append(row)
	else:
		# Header detection failed -> return raw rows so nothing is silently lost.
		for raw in rows[1:]:
			if any(cell not in (None, "") for cell in raw):
				extracted.append({"raw": [c for c in raw]})

	return extracted
