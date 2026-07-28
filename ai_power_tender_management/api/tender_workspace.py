# Copyright (c) 2026, milind and contributors
# For license information, please see license.txt
"""
Whitelisted API for the Tender Upload / Tender Workspace page (Phase 1 MVP).

This layer intentionally contains PLACEHOLDER logic for the "AI" steps:
  - analyze_tender_document() creates stub AI Summary rows.
  - extract_boq() does best-effort digital extraction, falling back to sample rows.

Everything is structured so the real AI / OCR pipeline can be dropped in later
without changing the frontend contract. Search for "TODO(AI)" markers below.
"""

import io
import hashlib
import json
import os
import re
from urllib.parse import quote

import frappe
from frappe import _
from frappe.utils import cint, flt
from frappe.utils.file_manager import save_file

from ai_power_tender_management.utils import ai_service, document_parser, orgchart, schedule

# Placeholder text shown until the real AI extraction is connected.
PLACEHOLDER_TEXT = "سيتم استبدال هذا النص بنتائج تحليل الذكاء الاصطناعي في المرحلة التالية."

# Document types that represent the main tender / specifications document.
TENDER_DOC_TYPES = ("Tender Document", "Terms and Specifications")
# Document types that represent the BOQ / pricing sheet.
BOQ_DOC_TYPES = ("BOQ", "Purchase Requisition")

# How many placeholder rows to create per summary type on analyze.
SUMMARY_BLUEPRINT = {
	"Tender Summary": 1,
	"Scope of Work": 1,
	"Important Requirement": 1,
	"Dangerous Clause": 3,
	"Missing Information": 2,
	"Technical Requirement": 1,
	"Commercial Condition": 1,
	"Submission Instruction": 1,
	"Warranty Requirement": 1,
	"Penalty Clause": 1,
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _get_doc(name):
	if not name:
		frappe.throw(_("Tender Workspace name is required."))
	return frappe.get_doc("Tender Workspace", name)


def _find_document(doc, document_types):
	"""Return the first uploaded document row matching one of document_types."""
	for row in doc.uploaded_documents:
		if row.document_type in document_types and row.file:
			return row
	return None


def _file_format_from_url(file_url):
	return (os.path.splitext(file_url or "")[1] or "").lstrip(".").lower()


_KNOWLEDGE_VERSION = "1"
_KNOWLEDGE_CHUNK_CHARS = 2500
_KNOWLEDGE_MAX_CHUNKS = 120
_RETRIEVED_CONTEXT_CHARS = 6500
_TENDER_INFO_TERMS = (
	"tender", "رقم", "منافسة", "اسم", "client", "الجهة", "closing", "إغلاق",
	"تاريخ", "ضمان", "bond", "estimated", "قيمة", "location", "المدينة",
)


def _has_field(doc, fieldname):
	return bool(getattr(doc, "meta", None) and doc.meta.has_field(fieldname))


def _set_if_field(doc, fieldname, value):
	if _has_field(doc, fieldname):
		doc.set(fieldname, value)


def _loads_json(value, default):
	if not value:
		return default
	try:
		data = json.loads(value)
	except Exception:
		return default
	return data if data is not None else default


def _dumps_json(value):
	return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


_LOG_PAYLOAD_CHARS = 1500


def _sanitize_log_text(value, max_chars=_LOG_PAYLOAD_CHARS):
	"""Redact obvious secrets and cap diagnostic snippets stored in Error Log."""
	if value is None:
		text = ""
	elif isinstance(value, str):
		text = value
	else:
		try:
			text = frappe.as_json(value)
		except Exception:
			text = str(value)

	for pattern in (
		r"sk-[A-Za-z0-9_\-]{10,}",
		r"(?i)(api[_-]?key\s*[:=]\s*)[^\s,\n]+",
		r"(?i)(authorization\s*[:=]\s*bearer\s+)[^\s,\n]+",
		r"(?i)(password\s*[:=]\s*)[^\s,\n]+",
	):
		text = re.sub(pattern, lambda m: (m.group(1) if m.groups() else "") + "[redacted]", text)

	if max_chars and len(text) > max_chars:
		return text[:max_chars] + "\n...[truncated]"
	return text


def _error_log_url(error_log):
	return f"/app/error-log/{quote(error_log)}" if error_log else None


def _remember_tender_error(tender_workspace_name, error_log):
	if not tender_workspace_name or not error_log:
		return
	logs = getattr(frappe.flags, "tender_workspace_error_logs", None) or {}
	logs[tender_workspace_name] = error_log
	frappe.flags.tender_workspace_error_logs = logs


def _latest_tender_error(tender_workspace_name):
	return (getattr(frappe.flags, "tender_workspace_error_logs", None) or {}).get(tender_workspace_name)


def _result_with_error_log(result, error_log):
	result = result or {}
	if error_log:
		result["error_log"] = error_log
		result["error_log_url"] = _error_log_url(error_log)
	return result


def _failure_result(status, message, log_title, tender_workspace_name, log_message=None, **extra):
	error_log = _log_tender_error(log_title, tender_workspace_name, message=log_message)
	result = {"status": status, "message": message}
	result.update(extra)
	return _result_with_error_log(result, error_log)


def _file_fingerprint(file_url):
	"""Cheap file-change fingerprint based on resolved path metadata."""
	path = document_parser._resolve_file_path(file_url)
	if path and os.path.exists(path):
		stat = os.stat(path)
		raw = f"{file_url}|{path}|{stat.st_size}|{stat.st_mtime_ns}"
	else:
		raw = file_url or ""
	return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _documents_fingerprint(doc):
	"""Fingerprint all uploaded files and mirror per-row values when possible."""
	parts = []
	for row in getattr(doc, "uploaded_documents", []) or []:
		if not getattr(row, "file", None):
			continue
		fingerprint = _file_fingerprint(row.file)
		if getattr(row, "meta", None) and row.meta.has_field("file_fingerprint"):
			row.file_fingerprint = fingerprint
		parts.append("|".join([
			row.name or "",
			row.document_type or "",
			row.file or "",
			fingerprint,
		]))
	return hashlib.sha256("\n".join(parts).encode("utf-8")).hexdigest()


def _clear_knowledge_cache_fields(doc):
	for fieldname in (
		"knowledge_version",
		"knowledge_fingerprint",
		"knowledge_updated_on",
		"knowledge_chunks_json",
		"structured_analysis_json",
		"boq_cache_json",
	):
		_set_if_field(doc, fieldname, None)


def _kb_text_filename(row_name):
	return f"tender-kb-text-{row_name}.txt"


def _read_kb_text_cache(tender_name, row_name):
	names = frappe.get_all(
		"File",
		filters={
			"attached_to_doctype": "Tender Workspace",
			"attached_to_name": tender_name,
			"file_name": _kb_text_filename(row_name),
		},
		pluck="name",
	)
	if not names:
		return ""
	try:
		content = frappe.get_doc("File", names[0]).get_content()
		if isinstance(content, bytes):
			content = content.decode("utf-8", errors="ignore")
		return content or ""
	except Exception:
		_log_tender_error(
			"Tender knowledge cache: read failed",
			tender_name,
			message=f"Row: {row_name}\nFile: {_kb_text_filename(row_name)}\n\n{frappe.get_traceback()}",
		)
		return ""


def _write_kb_text_cache(tender_name, row_name, text):
	filename = _kb_text_filename(row_name)
	existing = frappe.get_all(
		"File",
		filters={
			"attached_to_doctype": "Tender Workspace",
			"attached_to_name": tender_name,
			"file_name": filename,
		},
		pluck="name",
	)
	for name in existing:
		frappe.delete_doc("File", name, ignore_permissions=True, force=True)
	if (text or "").strip():
		save_file(filename, text.encode("utf-8"), "Tender Workspace", tender_name, is_private=1)


def _pdf_text_with_pages(file_url):
	pages = document_parser.extract_pdf_pages(file_url)
	return "\n\n".join(f"[Page {page_no}]\n{text.strip()}" for page_no, text in pages if text.strip())


def _source_text_for_row(doc, row, preferred_text_by_row=None):
	preferred_text_by_row = preferred_text_by_row or {}
	text = (
		preferred_text_by_row.get(row.name)
		or preferred_text_by_row.get(getattr(row, "file", None))
		or ""
	)
	if text.strip():
		return text

	cached = _read_kb_text_cache(doc.name, row.name)
	if cached.strip():
		return cached

	file_url = row.file
	file_format = (row.file_format or _file_format_from_url(file_url)).lower()
	if file_format == "pdf":
		text = _pdf_text_with_pages(file_url)
		if len(text.strip()) <= document_parser.READABLE_TEXT_THRESHOLD:
			text = _read_ocr_cache(doc.name, row.name)
	elif file_format in ("xlsx", "xls", "csv"):
		text = document_parser.excel_to_text_grid(file_url)
	else:
		text = ""
	return text or ""


_PAGE_MARKER_RE = re.compile(r"\[Page\s+(\d+)\]\s*", re.IGNORECASE)


def _page_blocks(text):
	matches = list(_PAGE_MARKER_RE.finditer(text or ""))
	if not matches:
		return [(None, text or "")]
	blocks = []
	for idx, match in enumerate(matches):
		start = match.end()
		end = matches[idx + 1].start() if idx + 1 < len(matches) else len(text)
		blocks.append((cint(match.group(1)), text[start:end]))
	return blocks


def _guess_section_label(text):
	value = (text or "").upper()
	if any(marker.upper() in value for marker in _BOQ_MARKERS):
		return "BOQ"
	low = (text or "").lower()
	for label, keywords in {
		"Submission": ("submission", "submit", "إغلاق", "تقديم", "موعد"),
		"Commercial": ("payment", "vat", "bond", "guarantee", "غرام", "ضمان", "دفعة"),
		"Technical": ("specification", "scope", "works", "requirement", "مواصفات", "نطاق", "اشتراط"),
		"Schedule": ("duration", "timeline", "milestone", "calendar", "مدة", "جدول"),
		"Organization": ("staff", "manpower", "engineer", "crew", "فريق", "مهندس", "عمالة"),
	}.items():
		if any(keyword in low for keyword in keywords):
			return label
	return "General"


def _chunk_source_text(source, text):
	chunks = []
	for page_no, body in _page_blocks(text):
		for chunk in _chunk_text(body, max_chars=_KNOWLEDGE_CHUNK_CHARS):
			chunk = chunk.strip()
			if not chunk:
				continue
			chunks.append({
				"id": f"{source['row_name']}:{len(chunks) + 1}",
				"source_row": source["row_name"],
				"source_document": source["source_document"],
				"document_type": source["document_type"],
				"file_format": source["file_format"],
				"page_start": page_no or "",
				"page_end": page_no or "",
				"section": _guess_section_label(chunk),
				"text": chunk,
			})
			if len(chunks) >= _KNOWLEDGE_MAX_CHUNKS:
				return chunks
	return chunks


def _boq_row_dict(row):
	return {
		"line_type": row.line_type or "Item",
		"item_no": row.item_no or "",
		"parent_item_no": row.parent_item_no or "",
		"description": row.description or "",
		"description_en": row.description_en or "",
		"unit": row.unit or "",
		"quantity": flt(row.quantity),
		"unit_price": flt(row.unit_price),
		"total": flt(row.total),
		"specification": row.specification or "",
		"source_page": row.source_page or "",
		"extraction_confidence": flt(row.extraction_confidence),
	}


def _structured_analysis_payload(doc, fingerprint):
	summaries = []
	for row in getattr(doc, "ai_summary", []) or []:
		if not getattr(row, "extracted_text", None):
			continue
		summaries.append({
			"summary_type": row.summary_type or "",
			"extracted_text": row.extracted_text or "",
			"source_document": row.source_document or "",
			"page_number": row.page_number or "",
			"confirmed": cint(row.confirmed),
		})

	tender_info = {
		"tender_name": doc.tender_name or "",
		"tender_name_ar": getattr(doc, "tender_name_ar", None) or "",
		"tender_number": doc.tender_number or "",
		"client_name": doc.client_name or "",
		"tender_type": doc.tender_type or "",
		"location": doc.location or "",
		"closing_date": str(doc.closing_date or ""),
		"closing_time": str(doc.closing_time or ""),
		"bid_bond_amount": flt(doc.bid_bond_amount),
		"estimated_value": flt(doc.estimated_value),
	}

	return {
		"version": _KNOWLEDGE_VERSION,
		"fingerprint": fingerprint,
		"updated_on": frappe.utils.now(),
		"tender_info": tender_info,
		"summaries": summaries,
		"evidence": [
			{
				"type": row["summary_type"],
				"source_document": row["source_document"],
				"page_number": row["page_number"],
				"text": row["extracted_text"][:500],
			}
			for row in summaries
		],
		"counts": {
			"summary_rows": len(summaries),
			"dangerous_clauses": sum(1 for row in summaries if row["summary_type"] == "Dangerous Clause"),
			"missing_information": sum(1 for row in summaries if row["summary_type"] == "Missing Information"),
		},
	}


def _boq_cache_payload(doc, fingerprint):
	rows = [_boq_row_dict(row) for row in getattr(doc, "boq_items", []) or [] if getattr(row, "description", None)]
	return {
		"version": _KNOWLEDGE_VERSION,
		"fingerprint": fingerprint,
		"updated_on": frappe.utils.now(),
		"rows": rows,
		"subtotal": sum(flt(row.get("total")) for row in rows),
		"items_count": len(rows),
	}


def _update_knowledge_cache_fields(doc, preferred_text_by_row=None, force_chunks=False):
	if not _has_field(doc, "knowledge_chunks_json"):
		return {"chunks": [], "fingerprint": ""}

	fingerprint = _documents_fingerprint(doc)
	cached_fingerprint = getattr(doc, "knowledge_fingerprint", None)
	chunks = _loads_json(getattr(doc, "knowledge_chunks_json", None), [])
	if force_chunks or cached_fingerprint != fingerprint or not chunks:
		chunks = []
		for row in getattr(doc, "uploaded_documents", []) or []:
			if not getattr(row, "file", None):
				continue
			file_format = (row.file_format or _file_format_from_url(row.file)).lower()
			source = {
				"row_name": row.name,
				"source_document": row.file_name or row.file,
				"document_type": row.document_type or "",
				"file_format": file_format,
			}
			text = _source_text_for_row(doc, row, preferred_text_by_row=preferred_text_by_row)
			if text.strip():
				_write_kb_text_cache(doc.name, row.name, text)
				chunks.extend(_chunk_source_text(source, text))
			if len(chunks) >= _KNOWLEDGE_MAX_CHUNKS:
				chunks = chunks[:_KNOWLEDGE_MAX_CHUNKS]
				break

	_set_if_field(doc, "knowledge_version", _KNOWLEDGE_VERSION)
	_set_if_field(doc, "knowledge_fingerprint", fingerprint)
	_set_if_field(doc, "knowledge_updated_on", frappe.utils.now())
	_set_if_field(doc, "knowledge_chunks_json", _dumps_json(chunks))
	_set_if_field(doc, "structured_analysis_json", _dumps_json(_structured_analysis_payload(doc, fingerprint)))
	_set_if_field(doc, "boq_cache_json", _dumps_json(_boq_cache_payload(doc, fingerprint)))
	return {"chunks": chunks, "fingerprint": fingerprint}


def _valid_cached_payload(doc, fieldname):
	fingerprint = _documents_fingerprint(doc)
	payload = _loads_json(getattr(doc, fieldname, None), {})
	if (
		isinstance(payload, dict)
		and payload.get("version") == _KNOWLEDGE_VERSION
		and payload.get("fingerprint") == fingerprint
	):
		return payload
	return {}


def _cached_tender_info(doc):
	info = _valid_cached_payload(doc, "structured_analysis_json").get("tender_info") or {}
	useful = ["tender_number", "client_name", "closing_date", "bid_bond_amount", "estimated_value"]
	return info if any(info.get(field) for field in useful) else None


def _cached_summary_rows(doc, source_document):
	payload = _valid_cached_payload(doc, "structured_analysis_json")
	rows = payload.get("summaries") or []
	matched = [row for row in rows if row.get("source_document") == source_document]
	return matched or None


def _cached_boq_rows(doc):
	payload = _valid_cached_payload(doc, "boq_cache_json")
	rows = payload.get("rows") or []
	return [_normalize_boq_row(row) for row in rows if isinstance(row, dict) and row.get("description")]


def _ensure_knowledge_base(doc):
	fingerprint = _documents_fingerprint(doc)
	chunks = _loads_json(getattr(doc, "knowledge_chunks_json", None), [])
	if (
		_has_field(doc, "knowledge_chunks_json")
		and getattr(doc, "knowledge_version", None) == _KNOWLEDGE_VERSION
		and getattr(doc, "knowledge_fingerprint", None) == fingerprint
		and chunks
	):
		return {"chunks": chunks, "fingerprint": fingerprint}
	if not getattr(doc, "name", None) or doc.is_new():
		return {"chunks": chunks, "fingerprint": fingerprint}
	payload = _update_knowledge_cache_fields(doc)
	doc.save(ignore_permissions=True)
	frappe.db.commit()
	return payload


def _query_terms(query):
	return {
		term.lower()
		for term in re.findall(r"[\w\u0600-\u06FF]{3,}", query or "")
	}


def _retrieved_knowledge_context(doc, query, max_chars=_RETRIEVED_CONTEXT_CHARS, limit=8):
	payload = _ensure_knowledge_base(doc)
	chunks = payload.get("chunks") or []
	if not chunks:
		return ""
	terms = _query_terms(query)
	if not terms:
		terms = {"scope", "requirement", "risk", "boq", "schedule", "manpower", "technical"}

	def score(chunk):
		value = f"{chunk.get('section', '')}\n{chunk.get('text', '')}".lower()
		hits = sum(value.count(term) for term in terms)
		if chunk.get("section") in ("BOQ", "Technical", "Commercial"):
			hits += 1
		return hits

	ranked = sorted(chunks, key=score, reverse=True)
	selected = [chunk for chunk in ranked if score(chunk) > 0][:limit] or ranked[: min(limit, len(ranked))]
	lines = []
	used = 0
	for chunk in selected:
		page = f" p.{chunk.get('page_start')}" if chunk.get("page_start") else ""
		prefix = f"- {chunk.get('source_document')} [{chunk.get('section')}{page}]: "
		text = re.sub(r"\s+", " ", chunk.get("text") or "").strip()
		remaining = max_chars - used - len(prefix) - 2
		if remaining <= 0:
			break
		body = text[:remaining]
		lines.append(prefix + body)
		used += len(prefix) + len(body) + 1
	return "\nRELEVANT SOURCE SNIPPETS:\n" + "\n".join(lines) if lines else ""


def _compact_text_for_terms(text, terms, max_chars=7000):
	text = (text or "").strip()
	if len(text) <= max_chars:
		return text
	terms = {term.lower() for term in terms if term}
	blocks = []
	for page_no, body in _page_blocks(text):
		for block in re.split(r"\n{2,}", body):
			block = block.strip()
			if block:
				blocks.append((page_no, block))

	def score(item):
		block = item[1].lower()
		return sum(block.count(term) for term in terms)

	ranked = sorted(blocks, key=score, reverse=True)
	selected = [item for item in ranked if score(item) > 0][:12]
	if blocks and blocks[0] not in selected:
		selected.append(blocks[0])
	out = []
	used = 0
	seen = set()
	for page_no, block in selected:
		key = block[:100]
		if key in seen:
			continue
		seen.add(key)
		prefix = f"[Page {page_no}]\n" if page_no else ""
		remaining = max_chars - used - len(prefix) - 2
		if remaining <= 0:
			break
		out.append(prefix + block[:remaining])
		used += len(out[-1]) + 2
	return "\n\n".join(out) or text[:max_chars]


# ---------------------------------------------------------------------------
# 1. Save (create or update) a Tender Workspace
# ---------------------------------------------------------------------------
@frappe.whitelist()
def save_tender_workspace(data):
	"""
	Create or update a Tender Workspace from basic info (+ optional documents).

	`data` is a JSON string / dict with the tender fields. If it contains a
	`name`, the existing record is updated; otherwise a new one is created.
	Optionally accepts a `documents` list to seed the uploaded_documents table.
	"""
	data = frappe.parse_json(data) if isinstance(data, str) else (data or {})

	name = data.get("name")
	if name and frappe.db.exists("Tender Workspace", name):
		doc = frappe.get_doc("Tender Workspace", name)
	else:
		doc = frappe.new_doc("Tender Workspace")

	basic_fields = [
		"tender_name", "tender_name_ar", "tender_number", "client_name", "portal_source",
		"closing_date", "reviewer", "status", "notes", "boq_currency",
		"vat_rate",
	]
	for field in basic_fields:
		if field in data:
			doc.set(field, data.get(field))

	if not doc.tender_name:
		frappe.throw(_("Tender Name is required."))

	if not doc.status:
		doc.status = "Draft"

	# Optional: seed uploaded document references (id-less rows appended).
	documents_added = False
	for d in (data.get("documents") or []):
		doc.append("uploaded_documents", {
			"document_type": d.get("document_type") or "Other Attachment",
			"file": d.get("file") or d.get("file_url"),
			"file_name": d.get("file_name"),
			"file_format": d.get("file_format") or _file_format_from_url(d.get("file") or d.get("file_url")),
			"ai_status": d.get("ai_status") or "Uploaded",
			"readable_status": d.get("readable_status") or "Unknown",
		})
		documents_added = True

	if documents_added:
		_clear_knowledge_cache_fields(doc)

	doc.save()
	frappe.db.commit()

	return {
		"name": doc.name,
		"status": doc.status,
		"message": _("Tender saved successfully"),
	}


# ---------------------------------------------------------------------------
# 2. Read a Tender Workspace (with child tables)
# ---------------------------------------------------------------------------
@frappe.whitelist()
def get_tender_workspace(name):
	"""Return the full Tender Workspace document including all child tables."""
	doc = _get_doc(name)
	return doc.as_dict()


# ---------------------------------------------------------------------------
# 3. Attach an uploaded file to the Tender Workspace
# ---------------------------------------------------------------------------
@frappe.whitelist()
def attach_tender_file(tender_workspace_name, document_type, file_url, file_name=None, file_format=None):
	"""
	Append an uploaded file to uploaded_documents and mark it as Uploaded.
	Returns the newly created document item row.
	"""
	doc = _get_doc(tender_workspace_name)

	row = doc.append("uploaded_documents", {
		"document_type": document_type or "Other Attachment",
		"file": file_url,
		"file_name": file_name or os.path.basename(file_url or ""),
		"file_format": (file_format or _file_format_from_url(file_url)).lower(),
		"ai_status": "Uploaded",
		"readable_status": "Unknown",
	})

	# Once at least one document is attached, reflect that in the tender status.
	if doc.status == "Draft":
		doc.status = "Documents Uploaded"

	_clear_knowledge_cache_fields(doc)
	doc.save()
	frappe.db.commit()

	return row.as_dict()


# ---------------------------------------------------------------------------
# 4. Analyze the tender document (PLACEHOLDER AI)
# ---------------------------------------------------------------------------
@frappe.whitelist()
def analyze_tender_document(tender_workspace_name):
	"""Queue analysis of the main Tender Document / Terms & Specifications file."""
	doc = _get_doc(tender_workspace_name)
	if not _find_document(doc, TENDER_DOC_TYPES):
		frappe.throw(_("Please upload Tender Document / Terms & Specifications first."))

	started = _enqueue_job(
		"analyze",
		doc.name,
		"ai_power_tender_management.api.tender_workspace._analyze_pipeline",
	)
	return {
		"status": "Processing",
		"background": True,
		"dangerous_clauses_count": 0,
		"missing_information_count": 0,
		"message": (
			_(
				"Analysing the tender document in the background. This can take a few minutes — "
				"the form will update automatically when it finishes."
			)
			if started
			else _("Analysis is already running in the background. Please wait for it to finish.")
		),
	}


def _analyze_pipeline(tender_workspace_name):
	"""Background: run document analysis and report the outcome to the form."""
	_run_pipeline(
		tender_workspace_name, "analyze", _("Reading the tender document…"), _analyze_tender_document_sync
	)


def _analyze_tender_document_sync(tender_workspace_name):
	"""
	Analyse the main Tender Document / Terms & Specifications file.

	Phase 1 behaviour:
	  - Digital PDFs: readable text -> create placeholder AI Summary rows.
	  - Non-readable / scanned PDFs: mark as "OCR Required" (no OCR performed).

	TODO(AI): replace the placeholder summary generation below with a call to
	the real LLM pipeline that consumes `document_parser.extract_text_from_pdf`.
	"""
	doc = _get_doc(tender_workspace_name)

	tender_doc = _find_document(doc, TENDER_DOC_TYPES)
	if not tender_doc:
		frappe.throw(_("Please upload Tender Document / Terms & Specifications first."))

	file_url = tender_doc.file
	file_format = (tender_doc.file_format or _file_format_from_url(file_url)).lower()
	source = tender_doc.file_name or tender_doc.file

	# Extract text once (PDF only). Non-PDF formats are treated as readable.
	document_text = document_parser.extract_text_from_pdf(file_url) if file_format == "pdf" else ""
	text_readable = (file_format != "pdf") or (
		len(document_text.strip()) > document_parser.READABLE_TEXT_THRESHOLD
	)

	ai_on = ai_service.is_enabled()
	cached_rows = _cached_summary_rows(doc, source)

	# No text layer + AI + OCR available -> OCR→chunk→AI pipeline (background job).
	# This is the preferred path for scanned PDFs: cheaper and rate-limit friendly.
	if (not cached_rows) and (not text_readable) and ai_on and file_format == "pdf" and document_parser.ocr_available():
		# Already inside the "analyze" background job, so run the OCR pipeline
		# inline — it publishes its own progress and its own final event.
		tender_doc.ai_status = "Processing"
		doc.save()
		frappe.db.commit()
		_ocr_analyze_pipeline(tender_workspace_name)
		return {"status": "Processed", "final_published": True}

	ai_rows = cached_rows
	vision_attempted = False

	if (not ai_rows) and text_readable and ai_on:
		# Readable digital text -> feed the text to the LLM.
		ai_rows = _ai_summary_rows(doc, document_text)
	elif (not ai_rows) and (not text_readable) and ai_on and file_format == "pdf" and ai_service.supports_pdf_vision():
		# No OCR available -> fall back to native PDF vision (may hit rate limits).
		vision_attempted = True
		ai_rows = _ai_summary_rows_pdf(doc, file_url)

	# Unreadable and no AI result -> OCR Required (logged for diagnosis).
	if (not text_readable) and not ai_rows:
		if vision_attempted:
			# The model was asked to read the PDF but failed (see the vision Error Log,
			# e.g. a rate-limit for a large document).
			log_title = "Tender Analyze: AI vision could not read PDF"
			user_msg = _(
				"AI could not read this PDF — it may be too large for the current AI rate limit "
				"(large documents must be processed in chunks). See the Error Log, then retry or raise the limit."
			)
		else:
			log_title = "Tender Analyze: OCR Required (no text layer)"
			user_msg = _(
				"This PDF has no readable text. Configure the AI Settings API key (vision) or use OCR."
			)
		error_log = _log_tender_error(
			log_title,
			doc.name,
			message=(
				f"File: {tender_doc.file_name} ({file_url})\n"
				f"Extracted text length: {len(document_text.strip())}\n"
				f"AI enabled: {ai_on} | PDF vision available: {ai_service.supports_pdf_vision()} | "
				f"vision attempted: {vision_attempted}\n\n"
				"The PDF has no extractable text layer (image/vector only). It needs AI vision "
				"(Anthropic) or OCR. Large PDFs may exceed the AI rate limit and must be chunked."
			),
		)
		tender_doc.ai_status = "OCR Required"
		tender_doc.readable_status = "OCR Required"
		tender_doc.ai_summary = user_msg
		doc.save()
		frappe.db.commit()
		return {
			"status": "OCR Required",
			"dangerous_clauses_count": 0,
			"missing_information_count": 0,
			"message": user_msg,
			"error_log": error_log,
			"error_log_url": _error_log_url(error_log),
		}

	# (Re)generate summary rows for this source (from text, vision, or placeholder).
	doc.ai_summary = [r for r in doc.ai_summary if r.source_document != source]

	if ai_rows:
		for row in ai_rows:
			doc.append("ai_summary", {
				"summary_type": row.get("summary_type"),
				"extracted_text": row.get("extracted_text") or "",
				"source_document": source,
				"page_number": str(row.get("page_number") or ""),
				"confirmed": 0,
			})
	else:
		for summary_type, count in SUMMARY_BLUEPRINT.items():
			for i in range(count):
				doc.append("ai_summary", {
					"summary_type": summary_type,
					"extracted_text": PLACEHOLDER_TEXT,
					"source_document": source,
					"page_number": str(i + 1),
					"confirmed": 0,
				})

	error_log = None
	if ai_rows:
		result_status = "Processed"
		result_message = _("Tender document analyzed successfully")
		child_summary = _("AI analysis complete.")
	elif not ai_on:
		result_status = "AI Not Configured"
		result_message = _("AI Settings API key is not configured. Placeholder analysis rows were created.")
		child_summary = _("Placeholder analysis created because AI Settings API key is not configured.")
		error_log = _log_tender_error(
			"Tender Analyze: AI not configured",
			doc.name,
			message=f"File: {tender_doc.file_name} ({file_url})\nReadable text length: {len(document_text.strip())}",
		)
	else:
		result_status = "AI Failed"
		result_message = _("AI could not extract tender-summary rows. Placeholder rows were created; open the Error Log below, then retry.")
		child_summary = _("Placeholder analysis created because AI returned no usable summary rows.")
		error_log = _log_tender_error(
			"Tender Analyze: no summary rows extracted",
			doc.name,
			message=(
				f"File: {tender_doc.file_name} ({file_url})\n"
				f"Format: {file_format}\nReadable text: {text_readable}\n"
				f"Readable text length: {len(document_text.strip())}\n"
				f"Vision attempted: {vision_attempted}"
			),
		)

	tender_doc.ai_status = "Processed" if ai_rows else "Failed"
	tender_doc.readable_status = "Yes"
	tender_doc.ai_summary = child_summary
	doc.status = "AI Analyzed"
	_update_knowledge_cache_fields(doc, preferred_text_by_row={tender_doc.name: document_text})
	doc.save()
	frappe.db.commit()

	dangerous = len([r for r in doc.ai_summary if r.summary_type == "Dangerous Clause"])
	missing = len([r for r in doc.ai_summary if r.summary_type == "Missing Information"])

	return _result_with_error_log({
		"status": result_status,
		"dangerous_clauses_count": dangerous,
		"missing_information_count": missing,
		"message": result_message,
	}, error_log)


# ---------------------------------------------------------------------------
# 4b. Extract tender header info (key fields) from the main document
# ---------------------------------------------------------------------------
# Key fields the AI auto-fills on the Tender Workspace from the document.
_TENDER_TYPES = ("Supply", "Works", "Services", "Consultancy", "Other")


@frappe.whitelist()
def extract_tender_info(tender_workspace_name):
	"""Queue extraction of the Tender Workspace key fields from the document."""
	doc = _get_doc(tender_workspace_name)
	if not _find_document(doc, TENDER_DOC_TYPES):
		frappe.throw(_("Please upload Tender Document / Terms & Specifications first."))

	started = _enqueue_job(
		"tender-info",
		doc.name,
		"ai_power_tender_management.api.tender_workspace._tender_info_pipeline",
	)
	return {
		"status": "Processing",
		"background": True,
		"filled": [],
		"message": (
			_("Extracting tender info in the background — the form will update when it finishes.")
			if started
			else _("Tender info is already being extracted in the background. Please wait.")
		),
	}


def _tender_info_pipeline(tender_workspace_name):
	"""Background: extract the tender header fields and report the outcome."""
	_run_pipeline(
		tender_workspace_name, "tender-info", _("Reading the tender document…"), _extract_tender_info_sync
	)


def _extract_tender_info_sync(tender_workspace_name):
	"""
	Auto-fill the Tender Workspace key fields (name, number, client, type,
	location, closing date/time, bid bond, estimated value) from the uploaded
	tender document using the LLM.

	Uses digital text when available, then cached OCR text, then native PDF
	vision. Only fields the model returns are overwritten; blanks are left alone.
	"""
	doc = _get_doc(tender_workspace_name)

	tender_doc = _find_document(doc, TENDER_DOC_TYPES)
	if not tender_doc:
		frappe.throw(_("Please upload Tender Document / Terms & Specifications first."))

	if not ai_service.is_enabled():
		return _failure_result(
			"AI Not Configured",
			_("AI Settings API key is not configured. Cannot extract tender info."),
			"Tender Extract Info: AI not configured",
			doc.name,
			log_message=f"File: {tender_doc.file_name} ({tender_doc.file})",
			filled=[],
		)

	file_url = tender_doc.file
	file_format = (tender_doc.file_format or _file_format_from_url(file_url)).lower()

	cached_info = _cached_tender_info(doc)
	if cached_info:
		filled = _apply_tender_info(doc, cached_info)
		_update_knowledge_cache_fields(doc)
		doc.save()
		frappe.db.commit()
		return {
			"status": "Extracted",
			"filled": filled,
			"message": _("Reused cached tender info ({0} field(s)).").format(len(filled)),
		}

	# 1) Digital text (or cached OCR text) -> text extraction.
	text = _best_tender_text(doc, tender_doc, file_format)
	info = _ai_tender_info(text) if text else None

	# 2) No usable text -> native PDF vision (Anthropic), when available.
	if not info and file_format == "pdf" and ai_service.supports_pdf_vision():
		info = _ai_tender_info_pdf(file_url)

	if not info:
		return _failure_result(
			"AI Failed",
			_("Could not extract tender info from the document. Open the Error Log below, then retry."),
			"Tender Extract Info: no fields extracted",
			doc.name,
			log_message=f"File: {tender_doc.file_name} ({file_url})\nFormat: {file_format}\nText length: {len((text or '').strip())}",
			filled=[],
		)

	filled = _apply_tender_info(doc, info)
	if not filled:
		return _failure_result(
			"AI Failed",
			_("AI found no usable tender fields to apply. Open the Error Log below, then retry."),
			"Tender Extract Info: no usable fields",
			doc.name,
			log_message=(
				f"File: {tender_doc.file_name} ({file_url})\n"
				f"Returned keys: {sorted(info.keys())}\n"
				f"Sanitized payload:\n{_sanitize_log_text(info)}"
			),
			filled=[],
		)
	_update_knowledge_cache_fields(doc, preferred_text_by_row={tender_doc.name: text})
	doc.save()
	frappe.db.commit()

	return {
		"status": "Extracted",
		"filled": filled,
		"message": _("Extracted {0} field(s) from the tender document.").format(len(filled)),
	}


def _best_tender_text(doc, tender_doc, file_format):
	"""Best available text for the tender doc: digital text, else cached OCR."""
	if file_format != "pdf":
		return ""
	text = _pdf_text_with_pages(tender_doc.file)
	if len(text.strip()) > document_parser.READABLE_TEXT_THRESHOLD:
		return text
	# Reuse OCR text cached by a previous analyze run, if any.
	return _read_ocr_cache(doc.name, tender_doc.name)


def _tender_info_schema():
	return (
		"{\"tender_name\": str, \"tender_name_ar\": Arabic tender name exactly as "
		"printed in the document (empty if the document is not Arabic), "
		"\"tender_number\": str, \"client_name\": str, "
		f"\"tender_type\": one of {list(_TENDER_TYPES)}, "
		"\"location\": city/region str, \"closing_date\": \"YYYY-MM-DD\" or empty, "
		"\"closing_time\": \"HH:MM\" 24h or empty, \"bid_bond_amount\": number or 0, "
		"\"estimated_value\": number or 0}"
	)


def _ai_tender_info(text):
	"""Ask the LLM to extract tender header fields from document text."""
	text = (text or "").strip()
	if len(text) < 100:
		return None
	system = (
		"You extract header/metadata fields from government tender documents "
		"(often Arabic). Respond ONLY with a valid JSON object, no prose."
	)
	prompt = (
		"From the tender document below, extract these fields as a single JSON "
		f"object: {_tender_info_schema()}. Use an empty string or 0 when a field "
		"is not present. Dates must be ISO format (YYYY-MM-DD).\n\n"
		f"DOCUMENT:\n{_compact_text_for_terms(text, _TENDER_INFO_TERMS)}"
	)
	data = ai_service.complete_json(prompt, system=system, task="extraction")
	return data if isinstance(data, dict) else None


def _ai_tender_info_pdf(file_url):
	"""Extract tender header fields by reading the PDF natively (vision)."""
	system = (
		"You extract header/metadata fields from a scanned Arabic government "
		"tender PDF. Respond ONLY with a valid JSON object, no prose."
	)
	prompt = (
		"Read the attached tender PDF and extract these fields as a single JSON "
		f"object: {_tender_info_schema()}. Use an empty string or 0 when a field "
		"is not present. Dates must be ISO format (YYYY-MM-DD)."
	)
	data = ai_service.complete_pdf_json(file_url, prompt, system=system, task="extraction")
	return data if isinstance(data, dict) else None


def _apply_tender_info(doc, info):
	"""Set doc fields from an AI info dict; return the list of fields filled."""
	filled = []

	def _set(field, value):
		if value in (None, "", 0, "0"):
			return
		doc.set(field, value)
		filled.append(field)

	_set("tender_name", str(info.get("tender_name") or "").strip())
	_set("tender_name_ar", str(info.get("tender_name_ar") or "").strip())
	_set("tender_number", str(info.get("tender_number") or "").strip())
	_set("client_name", str(info.get("client_name") or "").strip())
	_set("location", str(info.get("location") or "").strip())

	ttype = str(info.get("tender_type") or "").strip().title()
	if ttype in _TENDER_TYPES:
		_set("tender_type", ttype)

	if info.get("closing_date"):
		try:
			_set("closing_date", frappe.utils.getdate(info.get("closing_date")))
		except Exception:
			_log_tender_error(
				"Tender Extract Info: invalid closing_date",
				doc.name,
				message=f"Value: {_sanitize_log_text(info.get('closing_date'), 200)}",
			)
	if info.get("closing_time"):
		try:
			_set("closing_time", frappe.utils.get_time(str(info.get("closing_time"))))
		except Exception:
			_log_tender_error(
				"Tender Extract Info: invalid closing_time",
				doc.name,
				message=f"Value: {_sanitize_log_text(info.get('closing_time'), 200)}",
			)

	if flt(info.get("bid_bond_amount")):
		_set("bid_bond_amount", flt(info.get("bid_bond_amount")))
	if flt(info.get("estimated_value")):
		_set("estimated_value", flt(info.get("estimated_value")))

	return filled


# ---------------------------------------------------------------------------
# 5. Extract BOQ items
# ---------------------------------------------------------------------------
@frappe.whitelist()
def extract_boq(tender_workspace_name):
	"""Queue BOQ extraction from the uploaded BOQ / Purchase Requisition file."""
	doc = _get_doc(tender_workspace_name)
	if not (_find_document(doc, BOQ_DOC_TYPES) or _find_document(doc, TENDER_DOC_TYPES)):
		frappe.throw(_("Please upload a BOQ or Tender Document first."))

	started = _enqueue_job(
		"boq",
		doc.name,
		"ai_power_tender_management.api.tender_workspace._boq_pipeline",
	)
	return {
		"status": "Processing",
		"background": True,
		"items_count": 0,
		"message": (
			_(
				"Extracting the BOQ in the background. This can take a few minutes — the form "
				"will update automatically when it finishes."
			)
			if started
			else _("BOQ extraction is already running in the background. Please wait.")
		),
	}


def _boq_pipeline(tender_workspace_name):
	"""Background: run BOQ extraction and report the outcome to the form."""
	_run_pipeline(tender_workspace_name, "boq", _("Reading the BOQ…"), _extract_boq_sync)


def _extract_boq_sync(tender_workspace_name):
	"""
	Extract BOQ line items from the uploaded BOQ / Purchase Requisition file.

	Phase 1 behaviour:
	  - Excel (xls/xlsx/csv): best-effort row extraction via openpyxl.
	  - PDF: digital text check; if unreadable -> "OCR Required".
	  - If nothing reliable can be extracted -> sample placeholder rows.

	TODO(AI): replace the fallback sample rows with structured extraction from
	the AI pipeline once available.
	"""
	doc = _get_doc(tender_workspace_name)

	# BOQ is often embedded in the tender document (Saudi كراسة الشروط), so fall
	# back to the Tender Document when no dedicated BOQ file was uploaded.
	boq_doc = _find_document(doc, BOQ_DOC_TYPES) or _find_document(doc, TENDER_DOC_TYPES)
	if not boq_doc:
		frappe.throw(_("Please upload a BOQ or Tender Document first."))

	file_url = boq_doc.file
	file_format = (boq_doc.file_format or _file_format_from_url(file_url)).lower()

	rows = []
	status = "Extracted"
	ai_on = ai_service.is_enabled()
	source_text_for_cache = ""

	cached_rows = _cached_boq_rows(doc)
	if cached_rows:
		rows = cached_rows
		status = "Cached"

	if not rows and file_format in ("xlsx", "xls", "csv"):
		# Spreadsheet parsing is deterministic first. The LLM only handles
		# ambiguous layouts that column detection cannot map.
		rows = _rows_from_excel_deterministic(file_url)
		if not rows and ai_on:
			source_text_for_cache = document_parser.excel_to_text_grid(file_url)
			rows, _confidence, _issues = _extract_boq_from_text(source_text_for_cache, source_label=boq_doc.file_name)
	elif not rows and file_format == "pdf":
		boq_text = document_parser.extract_text_from_pdf(file_url)
		source_text_for_cache = boq_text
		text_readable = len(boq_text.strip()) > document_parser.READABLE_TEXT_THRESHOLD

		if text_readable and ai_on:
			# Digital PDF -> two-pass extraction from the text.
			rows, _confidence, _issues = _extract_boq_from_text(boq_text, source_label=boq_doc.file_name)
		elif (not text_readable) and ai_on and document_parser.ocr_available():
			# No text layer -> OCR→AI BOQ extraction. Already inside the "boq"
			# background job, so run it inline; it publishes its own final event.
			boq_doc.ai_status = "Processing"
			doc.save()
			frappe.db.commit()
			_ocr_boq_pipeline(tender_workspace_name)
			return {"status": "Extracted", "final_published": True}
		elif (not text_readable) and ai_service.supports_pdf_vision():
			# No OCR available -> read the PDF natively (vision).
			ai_rows = _ai_boq_rows_pdf(file_url)
			if ai_rows:
				rows = ai_rows

			# Still nothing and the PDF has no text layer -> OCR Required (logged).
			if not rows and not text_readable:
				error_log = _log_tender_error(
					"Tender Extract BOQ: OCR Required (no text layer)",
					doc.name,
					message=(
						f"File: {boq_doc.file_name} ({file_url})\n"
						f"Extracted text length: {len(boq_text.strip())}\n"
						f"AI enabled: {ai_on}\n"
						f"PDF vision available: {ai_service.supports_pdf_vision()}\n\n"
						"The BOQ PDF has no extractable text layer. Configure AI vision or OCR."
					),
				)
				boq_doc.ai_status = "OCR Required"
				boq_doc.readable_status = "OCR Required"
				doc.save()
				frappe.db.commit()
				return _result_with_error_log({
					"status": "OCR Required",
					"items_count": 0,
					"message": _("This BOQ PDF has no readable text. Configure the AI Settings API key (vision) or use OCR."),
				}, error_log)

	# Excel gave nothing but AI is on with a digital-text PDF handled above; nothing more to try here.

	# Items-only: drop title/subtotal/VAT/total footer headings (keep headings
	# that are referenced as a parent so links never orphan).
	rows = _filter_boq_rows(rows)

	# Nothing reliably extracted -> fail honestly. Never fabricate placeholder
	# items or prices: an empty/ambiguous BOQ must not look like a real one.
	if not rows:
		doc.set("boq_items", [])
		boq_doc.ai_status = "Failed"
		doc.save()
		frappe.db.commit()
		error_log = _log_tender_error(
			"Tender Extract BOQ: no items extracted",
			doc.name,
			message=(
				f"File: {boq_doc.file_name} ({file_url})\n"
				f"Format: {file_format}\nAI enabled: {ai_on}\n"
				"No BOQ line items could be extracted (unrecognised columns / empty "
				"result). No placeholder rows were inserted."
			),
		)
		return _result_with_error_log({
			"status": "No Items Found",
			"items_count": 0,
			"message": _(
				"No BOQ items could be extracted from this file. Please check that it "
				"contains a recognisable Bill of Quantities table (item / description / "
				"unit / quantity columns)."
			),
		}, error_log)

	# Replace existing BOQ items with the freshly extracted set.
	doc.set("boq_items", [])
	for r in rows:
		doc.append("boq_items", r)

	boq_doc.ai_status = "Extracted"
	boq_doc.readable_status = boq_doc.readable_status if boq_doc.readable_status == "OCR Required" else "Yes"
	doc.status = "BOQ Extracted"
	_update_knowledge_cache_fields(doc, preferred_text_by_row={boq_doc.name: source_text_for_cache})
	doc.save()
	frappe.db.commit()

	return {
		"status": status,
		"items_count": len(doc.boq_items),
		"message": _("BOQ extracted successfully"),
	}


# ---------------------------------------------------------------------------
# 7. Generate proposal sections
# ---------------------------------------------------------------------------
# Ordered list of proposal sections generated for every tender.
PROPOSAL_SECTIONS = [
	"Scope Understanding",
	"Methodology",
	"Implementation Plan",
	"Project Timeline",
	"Equipment List",
	"Organization Chart",
	"QA/QC Plan",
	"HSE Plan",
	"Compliance Matrix",
	"Risk Summary",
]
# The one section rendered from structured schedule rows instead of prose.
SCHEDULE_SECTION = "Project Timeline"

# Legacy section names accepted from older clients; stored rows are migrated by
# `patches.rename_primavera_timeline_section`.
PROPOSAL_SECTION_ALIASES = {}

# Arabic display titles for the proposal sections. `section_type` stays English
# (it is the canonical key used for ordering, aliases and guidance lookups); only
# the shown `title` is Arabic so the proposal reads natively RTL.
PROPOSAL_SECTION_TITLES_AR = {
	"Scope Understanding": "فهم نطاق العمل",
	"Methodology": "منهجية العمل",
	"Implementation Plan": "خطة التنفيذ",
	"Project Timeline": "الجدول الزمني للمشروع",
	"Equipment List": "قائمة المعدات",
	"Organization Chart": "الهيكل التنظيمي",
	"QA/QC Plan": "خطة ضمان ومراقبة الجودة",
	"HSE Plan": "خطة الصحة والسلامة والبيئة",
	"Compliance Matrix": "مصفوفة الامتثال",
	"Risk Summary": "ملخص المخاطر",
}


def _proposal_title(section):
	"""Arabic display title for a proposal section (falls back to the key)."""
	return PROPOSAL_SECTION_TITLES_AR.get(section, section)

# Arabic placeholder content shown until the real AI generation is connected.
PLACEHOLDER_PROPOSAL_CONTENT = "سيتم إنشاء محتوى هذا القسم بواسطة الذكاء الاصطناعي في المرحلة التالية."

# Generate proposal text in small batches. Asking for all sections in one
# compact JSON response makes models compress each section too aggressively.
_PROPOSAL_BATCH_SIZE = 3
_PROPOSAL_BATCH_MAX_TOKENS = 3500
_PROPOSAL_SECTION_WORD_TARGET = "220-350"
PROPOSAL_SECTION_GUIDANCE = {
	"Scope Understanding": "project objective, scope boundaries, assumptions, deliverables, and client priorities",
	"Methodology": "execution approach, coordination, approvals, procurement, installation, testing, and handover",
	"Implementation Plan": "phases, responsibilities, dependencies, sequence of activities, and control points",
	"Project Timeline": "WBS-style milestones, activity durations, dependencies, and progress monitoring",
	"Equipment List": "equipment, tools, instruments, manpower resources, and mobilization readiness",
	"Organization Chart": "project governance, reporting lines, key roles, responsibilities, and escalation path",
	"QA/QC Plan": "ITP, inspections, material approvals, testing, nonconformance control, and records",
	"HSE Plan": "risk assessment, permits, PPE, emergency response, toolbox talks, and site housekeeping",
	"Compliance Matrix": "requirement-by-requirement compliance approach, evidence, responsibility, and remarks",
	"Risk Summary": "major technical, commercial, schedule, contractual, HSE, and compliance risks with mitigation actions",
}


def _normalize_proposal_section(section_type):
	"""Return the canonical proposal section name, accepting known UI aliases."""
	section = (section_type or "").strip()
	return PROPOSAL_SECTION_ALIASES.get(section, section)


def _order_proposal_sections(doc):
	"""Keep proposal child rows in the same order as the exported proposal."""
	order = {section: idx for idx, section in enumerate(PROPOSAL_SECTIONS)}
	doc.proposal_sections.sort(key=lambda row: order.get(row.section_type, len(order)))
	for idx, row in enumerate(doc.proposal_sections, start=1):
		row.idx = idx


@frappe.whitelist()
def generate_proposal_sections(tender_workspace_name):
	"""
	Queue (re)generation of all standard proposal sections for the tender.

	Writing ten detailed Arabic sections takes several LLM round-trips (minutes),
	so the work runs in a background job; the form shows live progress in its
	"Background Processes" panel and reloads itself when the job finishes.
	"""
	doc = _get_doc(tender_workspace_name)

	started = _enqueue_job(
		"proposal",
		doc.name,
		"ai_power_tender_management.api.tender_workspace._proposal_sections_pipeline",
	)
	if not started:
		return {
			"status": "Processing",
			"background": True,
			"message": _("Proposal sections are already being generated in the background."),
		}

	return {
		"status": "Processing",
		"background": True,
		"message": _(
			"Generating proposal sections in the background. This can take a few minutes — "
			"the form will update automatically when it finishes."
		),
	}


def _proposal_sections_pipeline(tender_workspace_name):
	"""Background: generate every proposal section, publishing progress per batch."""
	_run_pipeline(
		tender_workspace_name,
		"proposal",
		_("Preparing tender context…"),
		_generate_proposal_sections_sync,
	)


def _generate_proposal_sections_sync(tender_workspace_name):
	"""
	(Re)generate the standard proposal sections for the tender.

	When AI Settings are enabled, create detailed Arabic proposal text using
	the AI summary and BOQ context. Otherwise, create the same rows with
	placeholder Arabic content. Sets the tender status to "Proposal Drafted".
	"""
	doc = _get_doc(tender_workspace_name)

	# Ask the LLM (from AI Settings) for Arabic content per section when enabled.
	ai_enabled = ai_service.is_enabled()

	def on_progress(done, total):
		# 10% for setup, 80% for generation, the rest for the save below.
		_publish(
			doc.name,
			_("Wrote {0} of {1} proposal sections…").format(done, total),
			10 + int(80 * done / max(1, total)),
			job_key="proposal",
		)

	ai_map = _ai_proposal_map(doc, on_progress=on_progress) if ai_enabled else None

	# Reload before mutating the child table: the AI loop above runs for minutes
	# and the user may have edited the tender meanwhile (avoids a stale save).
	doc = frappe.get_doc("Tender Workspace", tender_workspace_name)

	# Regenerate from scratch so re-running is idempotent.
	doc.set("proposal_sections", [])
	generated_count = 0
	for section in PROPOSAL_SECTIONS:
		content = (ai_map or {}).get(section)
		generated = bool(content)
		if generated:
			generated_count += 1
		content = content or PLACEHOLDER_PROPOSAL_CONTENT
		doc.append("proposal_sections", {
			"section_type": section,
			"title": _proposal_title(section),
			"status": "Generated" if generated else "Not Generated",
			"content": content,
			"confirmed": 0,
		})
	_order_proposal_sections(doc)

	doc.status = "Proposal Drafted"
	doc.save()
	frappe.db.commit()

	if generated_count == len(PROPOSAL_SECTIONS):
		status = "Generated"
		message = _("Proposal sections generated successfully")
		error_log = None
	elif not ai_enabled:
		status = "AI Not Configured"
		message = _("AI Settings API key is not configured. Placeholder proposal rows were created.")
		error_log = _log_tender_error(
			"Tender Proposal: AI not configured",
			doc.name,
			message=f"Requested sections: {', '.join(PROPOSAL_SECTIONS)}",
		)
	elif generated_count:
		status = "Partial"
		message = _("Generated {0} of {1} proposal sections. Missing sections use placeholders; check Error Log.").format(
			generated_count, len(PROPOSAL_SECTIONS)
		)
		error_log = _latest_tender_error(doc.name) or _log_tender_error(
			"Tender Proposal: partial generation",
			doc.name,
			message=f"Generated {generated_count} of {len(PROPOSAL_SECTIONS)} sections.",
		)
	else:
		status = "AI Failed"
		message = _("AI could not generate proposal sections. Placeholder rows were created; check Error Log.")
		error_log = _latest_tender_error(doc.name) or _log_tender_error(
			"Tender Proposal: no sections generated",
			doc.name,
			message=f"Requested sections: {', '.join(PROPOSAL_SECTIONS)}",
		)

	return _result_with_error_log({
		"status": status,
		"sections_count": len(doc.proposal_sections),
		"generated_count": generated_count,
		"message": message,
	}, error_log)


@frappe.whitelist(methods=["POST"])
def generate_proposal_section(tender_workspace_name: str, section_type: str):
	"""Queue generation of one proposal section without touching the others."""
	doc = _get_doc(tender_workspace_name)
	section = _normalize_proposal_section(section_type)
	if section not in PROPOSAL_SECTIONS:
		frappe.throw(_("Unknown proposal section: {0}").format(section_type))

	started = _enqueue_job(
		"proposal-section",
		doc.name,
		"ai_power_tender_management.api.tender_workspace._proposal_section_pipeline",
		section_type=section,
	)
	if not started:
		return {
			"status": "Processing",
			"background": True,
			"section_type": section,
			"message": _("A proposal section is already being generated. Please wait for it to finish."),
		}

	return {
		"status": "Processing",
		"background": True,
		"section_type": section,
		"message": _("Generating {0} in the background…").format(_(section)),
	}


def _proposal_section_pipeline(tender_workspace_name, section_type):
	"""Background: generate a single proposal section."""
	_run_pipeline(
		tender_workspace_name,
		"proposal-section",
		_("Generating {0}…").format(_(section_type)),
		_generate_proposal_section_sync,
		section_type,
	)


def _generate_proposal_section_sync(tender_workspace_name, section_type):
	"""Generate or refresh one proposal section row without touching the others."""
	doc = _get_doc(tender_workspace_name)
	section = _normalize_proposal_section(section_type)
	if section not in PROPOSAL_SECTIONS:
		frappe.throw(_("Unknown proposal section: {0}").format(section_type))

	ai_enabled = ai_service.is_enabled()
	ai_map = _ai_proposal_map(doc, sections=[section]) if ai_enabled else None
	content = (ai_map or {}).get(section)
	generated = bool(content)
	content = content or PLACEHOLDER_PROPOSAL_CONTENT

	# Reload before mutating: the AI call above can take minutes.
	doc = frappe.get_doc("Tender Workspace", tender_workspace_name)

	target = None
	for row in doc.proposal_sections:
		if row.section_type == section:
			target = row
			break
	if not target:
		target = doc.append("proposal_sections", {})

	target.section_type = section
	target.title = _proposal_title(section)
	target.status = "Generated" if generated else "Not Generated"
	target.content = content
	target.confirmed = 0

	_order_proposal_sections(doc)
	doc.status = "Proposal Drafted"
	doc.save()
	frappe.db.commit()

	if generated:
		status = "Generated"
		message = _("Proposal section generated: {0}").format(section)
		error_log = None
	elif not ai_enabled:
		status = "AI Not Configured"
		message = _("AI Settings API key is not configured. Placeholder was created for: {0}").format(section)
		error_log = _log_tender_error(
			"Tender Proposal Section: AI not configured",
			doc.name,
			message=f"Section: {section}",
		)
	else:
		status = "AI Failed"
		message = _("AI could not generate {0}. Placeholder was created; check Error Log.").format(section)
		error_log = _latest_tender_error(doc.name) or _log_tender_error(
			"Tender Proposal Section: no content generated",
			doc.name,
			message=f"Section: {section}",
		)

	return _result_with_error_log({
		"status": status,
		"section_type": section,
		"generated": generated,
		"message": message,
	}, error_log)


# ---------------------------------------------------------------------------
# 7b. Baseline schedule
#   The schedule is generated as structured rows (not prose) so it can be
#   rendered as the activity table and Gantt chart the client expects. The AI
#   supplies the dates, float and critical flags; we rescale the weights and run
#   consistency checks over what it returns.
# ---------------------------------------------------------------------------
_SCHEDULE_MAX_TOKENS = 4000

# The shape of one activity, restated in the prompt so the model has no room to
# improvise field names.
_SCHEDULE_ROW_SCHEMA = (
	'{"wbs": "sub-group code inside this phase, e.g. 2.1 / 2.2 — SEVERAL activities '
	'SHARE the same code; never give every activity its own", '
	'"activity_id": "unique code like PRJ-ENG-010", '
	'"activity_name": "Arabic activity name", '
	'"activity_type": "Task Dependent" | "Start Milestone" | "Finish Milestone" | "Level of Effort", '
	'"original_duration": integer working days (0 for milestones), '
	'"predecessors": "comma-separated links like PRJ-ENG-010FS,PRJ-ENG-020SS+5" (empty for the first activity), '
	'"weight_pct": number, '
	'"primary_resource": "Arabic resource/crew name"}'
)

_SCHEDULE_RULES = (
	"RULES:\n"
	"- Do NOT supply dates, float or a critical flag. Those are computed from "
	"your durations and predecessor links, so only the network must be right.\n"
	"- Durations are in working days on a 6-day week (Saturday to Thursday).\n"
	"- Every activity except the very first must reference at least one "
	"predecessor by its exact activity_id, using FS/SS/FF relations with an "
	"optional +lag in working days. Accurate links matter more than anything "
	"else here: they determine the whole schedule.\n"
	"- BUILD A NETWORK, NOT A CHAIN. Do not simply link each activity to the one "
	"before it. Work that can proceed independently MUST run in parallel: give "
	"several activities the SAME predecessor so they start together, and link an "
	"activity to EVERY input it truly waits on, not just the most recent one. In "
	"a real baseline only about a quarter of activities end up on the critical "
	"path — if almost everything chains one-to-one, the schedule is wrong.\n"
	"- Use SS+lag where work overlaps (a follow-on trade starting a few days "
	"into its predecessor) and FF where two activities must finish together. "
	"Do not make every link FS.\n"
	"- Never create a circular dependency.\n"
	"- A Level of Effort activity (supervision, daily reporting) must ALSO be "
	"linked forward to the milestone that closes its phase, using FF. Without "
	"that it dangles and its float becomes the whole remaining project.\n"
	"- weight_pct reflects share of project cost/effort, not duration: short but "
	"expensive works (paving, racking, MEP) weigh far more than long supervision "
	"activities. Values are rescaled afterwards, so they need not total 100.\n"
	"- Open with a commencement milestone and close with a handover milestone, "
	"both zero duration.\n"
	"- activity_name and primary_resource must be in Arabic. All other fields ASCII.\n"
	"- Divide the phase into 2-4 WBS sub-groups only. Every activity in a sub-group "
	"carries that sub-group's code, so codes repeat down the rows."
)

_SCHEDULE_PHASE_COUNT = "9-12"
_SCHEDULE_PHASE_ACTIVITIES = "6-12"


def _as_object_list(data):
	"""
	Coerce the model's reply to a list of dicts.

	A bare JSON array is what the prompt asks for, but the model often wraps it
	as {"phases": [...]} or {"activities": [...]}. Requiring a bare list made
	that a silent, unlogged failure.
	"""
	if isinstance(data, list):
		return [item for item in data if isinstance(item, dict)]
	if isinstance(data, dict):
		# A single wrapped array — take the longest list of objects it holds.
		candidates = [
			[item for item in value if isinstance(item, dict)]
			for value in data.values()
			if isinstance(value, list)
		]
		candidates = [c for c in candidates if c]
		if candidates:
			return max(candidates, key=len)
		# Or a single object that is itself one record.
		if data:
			return [data]
	return []


def _schedule_start_date(doc):
	"""Assumed project start: shortly after the tender closes."""
	base = doc.closing_date or frappe.utils.nowdate()
	return frappe.utils.add_days(frappe.utils.getdate(base), 30)


def _ai_schedule_phases(doc, context):
	"""Ask for the WBS breakdown first — a small response that frames the rest."""
	system = (
		"You are a senior planning engineer preparing a Primavera-style baseline "
		"schedule for a Saudi government/utility tender. Respond ONLY with valid JSON."
	)
	prompt = (
		f"Break this project into {_SCHEDULE_PHASE_COUNT} sequential WBS phases, in "
		"execution order (mobilisation and design first, testing and handover last). "
		'Return a JSON array of {"wbs": "1.0", "title_ar": "Arabic phase name", '
		'"id_prefix": "3-5 uppercase letters for activity ids in this phase"}.\n\n'
		f"TENDER CONTEXT:\n{context}"
	)
	data = ai_service.complete_json(prompt, system=system, max_tokens=1500, task="extraction")
	items = _as_object_list(data)
	if not items:
		_log_tender_error(
			"Tender AI: no WBS phases in reply", doc.name,
			message=f"Sanitized payload:\n{_sanitize_log_text(data)}",
		)
		return None
	phases = []
	for item in items:
		wbs = str(item.get("wbs") or "").strip()
		title = str(item.get("title_ar") or "").strip()
		if wbs and title:
			phases.append({
				"wbs": wbs,
				"title_ar": title,
				"id_prefix": (str(item.get("id_prefix") or "ACT").strip() or "ACT")[:6].upper(),
			})
	return phases or None


def _schedule_rows_digest(rows):
	"""Compact list of already-planned activities, for cross-phase linking.

	Ids and names only — there are no dates to show, since the whole network is
	scheduled once at the end.
	"""
	if not rows:
		return "(none yet — this is the first phase)"
	return "\n".join(f"- {r.get('activity_id')} | {r.get('activity_name')}" for r in rows)


def _ai_schedule_phase_rows(doc, context, phase, done_rows, start_date):
	"""Generate one phase's activities, linked to everything planned so far."""
	system = (
		"You are a senior planning engineer building a Primavera-style baseline "
		"schedule. Respond ONLY with a valid JSON array, no prose."
	)
	prompt = (
		f"Plan WBS phase {phase['wbs']} — {phase['title_ar']} — as "
		f"{_SCHEDULE_PHASE_ACTIVITIES} activities.\n\n"
		f"Use activity ids of the form {phase['id_prefix']}-010, {phase['id_prefix']}-020, "
		"incrementing by ten.\n"
		f"The project starts on {start_date}.\n\n"
		f"Return a JSON array where each element is:\n{_SCHEDULE_ROW_SCHEMA}\n\n"
		f"{_SCHEDULE_RULES}\n\n"
		"ACTIVITIES ALREADY PLANNED IN EARLIER PHASES (link to these by id where "
		f"this phase depends on them):\n{_schedule_rows_digest(done_rows)}\n\n"
		f"TENDER CONTEXT:\n{context}"
	)
	data = ai_service.complete_json(prompt, system=system, max_tokens=_SCHEDULE_MAX_TOKENS, task="extraction")
	items = _as_object_list(data)
	if not items:
		_log_tender_error(
			f"Tender AI: no activities returned for WBS {phase['wbs']}", doc.name,
			message=f"Sanitized payload:\n{_sanitize_log_text(data)}",
		)
		return []
	return [row for row in (_normalize_schedule_row(item, phase) for item in items) if row]


_SCHEDULE_ACTIVITY_TYPES = (
	"Task Dependent", "Start Milestone", "Finish Milestone", "Level of Effort",
)


def _coerce_wbs(value, phase_wbs):
	"""
	Keep a row's WBS inside its phase.

	The model is asked for a handful of shared sub-group codes per phase, but it
	sometimes numbers every activity separately (1.1, 1.2, … 1.12) or wanders
	into another phase's numbering. Anything whose major number does not match
	the phase falls back to the phase code.
	"""
	major = str(phase_wbs or "").split(".")[0].strip()
	candidate = str(value or "").strip()
	if re.fullmatch(r"\d+(\.\d+)?", candidate) and candidate.split(".")[0] == major:
		return candidate[:20]
	return str(phase_wbs or "").strip()[:20]


def _normalize_schedule_row(item, phase):
	"""Coerce one AI activity dict into a safe child-row dict, or None."""
	activity_id = str(item.get("activity_id") or "").strip()
	if not activity_id:
		return None

	activity_type = str(item.get("activity_type") or "").strip()
	if activity_type not in _SCHEDULE_ACTIVITY_TYPES:
		activity_type = "Task Dependent"

	# Milestones are zero-duration by definition; everything else takes at least
	# a day, otherwise it collapses to a point on the chart.
	duration = max(0, frappe.utils.cint(item.get("original_duration")))
	if activity_type in ("Start Milestone", "Finish Milestone"):
		duration = 0
	elif duration == 0:
		duration = 1

	# Truncation limits mirror the DocType field lengths. Trimming here rather
	# than at save() matters: the save happens only after every phase has been
	# generated, so one over-long name would otherwise discard the whole run.
	return {
		"wbs": _coerce_wbs(item.get("wbs"), phase["wbs"]),
		"activity_id": activity_id[:60],
		"activity_name": str(item.get("activity_name") or "").strip()[:300],
		"activity_type": activity_type,
		"calendar": "6D-8H",
		"original_duration": duration,
		# planned_start / planned_finish / total_float / is_critical are left
		# unset here — schedule.apply_cpm derives them from the network.
		"predecessors": str(item.get("predecessors") or "").strip()[:500],
		"weight_pct": max(0.0, frappe.utils.flt(item.get("weight_pct"))),
		"primary_resource": str(item.get("primary_resource") or "").strip()[:140],
		"status": "Not Started",
	}


@frappe.whitelist()
def generate_schedule(tender_workspace_name):
	"""
	Queue (re)generation of the baseline schedule for the tender.

	Around a hundred activities take several LLM round-trips, so the work runs
	in a background job like the other AI steps.
	"""
	doc = _get_doc(tender_workspace_name)

	if not ai_service.is_enabled():
		return _failure_result(
			"AI Not Configured",
			_("AI Settings API key is not configured, so no schedule can be generated."),
			"Tender Schedule: AI not configured",
			doc.name,
		)

	started = _enqueue_job(
		"schedule",
		doc.name,
		"ai_power_tender_management.api.tender_workspace._schedule_pipeline",
	)
	if not started:
		return {
			"status": "Processing",
			"background": True,
			"message": _("The schedule is already being generated in the background."),
		}

	return {
		"status": "Processing",
		"background": True,
		"message": _(
			"Generating the baseline schedule in the background. This can take a few "
			"minutes — the form will update automatically when it finishes."
		),
	}


def _schedule_pipeline(tender_workspace_name):
	"""Background: build the schedule, publishing progress per WBS phase."""
	try:
		_publish(tender_workspace_name, _("Preparing tender context…"), 5, job_key="schedule")
		result = _generate_schedule_sync(tender_workspace_name)
		state = _job_state_for_result(result)
		error_log = result.get("error_log") or (
			_latest_tender_error(tender_workspace_name) if state in ("failed", "warning") else None
		)
		_publish(
			tender_workspace_name,
			result.get("message") or _("Schedule generated."),
			100,
			reload=True,
			job_key="schedule",
			state=state,
			error_log=error_log,
		)
	except Exception:
		frappe.db.rollback()
		log = _log_tender_error("Tender schedule generation failed", tender_workspace_name)
		_publish(
			tender_workspace_name,
			_("Schedule generation failed."),
			100,
			reload=True,
			job_key="schedule",
			state="failed",
			error_log=log,
		)


def _generate_schedule_sync(tender_workspace_name):
	"""(Re)generate `schedule_activities` phase by phase, then normalise them."""
	doc = _get_doc(tender_workspace_name)

	if not ai_service.is_enabled():
		return _failure_result(
			"AI Not Configured",
			_("AI Settings API key is not configured, so no schedule can be generated."),
			"Tender Schedule: AI not configured",
			doc.name,
		)

	context = _proposal_generation_context(
		doc,
		query="baseline schedule timeline phases milestones durations dependencies BOQ scope technical requirements",
	)
	start_date = frappe.utils.formatdate(_schedule_start_date(doc), "yyyy-MM-dd")

	_publish(doc.name, _("Planning the work breakdown structure…"), 10, job_key="schedule")
	phases = _ai_schedule_phases(doc, context)
	if not phases:
		error_log = _latest_tender_error(doc.name) or _log_tender_error(
			"Tender Schedule: no WBS phases generated",
			doc.name,
			message=f"Context length: {len(context or '')}",
		)
		return _result_with_error_log({
			"status": "AI Failed",
			"message": _("AI could not plan the work breakdown structure; check the Error Log."),
		}, error_log)

	rows = []
	for idx, phase in enumerate(phases, start=1):
		_publish(
			doc.name,
			_("Planning phase {0} of {1}: {2}").format(idx, len(phases), phase["title_ar"]),
			10 + int(80 * (idx - 1) / len(phases)),
			job_key="schedule",
		)
		rows.extend(_ai_schedule_phase_rows(doc, context, phase, rows, start_date))

	if not rows:
		error_log = _latest_tender_error(doc.name) or _log_tender_error(
			"Tender Schedule: no activities generated",
			doc.name,
			message=f"Phase count: {len(phases)}",
		)
		return _result_with_error_log({
			"status": "AI Failed",
			"message": _("AI returned no schedule activities; check the Error Log."),
		}, error_log)

	# Reload before mutating: the loop above runs for minutes and the user may
	# have edited the tender meanwhile.
	doc = frappe.get_doc("Tender Workspace", tender_workspace_name)
	doc.set("schedule_activities", [])
	for row in rows:
		doc.append("schedule_activities", row)

	# Dates, float and the critical path are computed from the network the AI
	# described, not taken from it.
	_publish(doc.name, _("Calculating dates and critical path…"), 92, job_key="schedule")
	try:
		info = schedule.apply_cpm(doc.schedule_activities, _schedule_start_date(doc))
	except schedule.cpm.ScheduleError:
		error_log = _log_tender_error("Tender schedule: CPM could not schedule the network", doc.name)
		return _result_with_error_log({
			"status": "AI Failed",
			"message": _("The generated activities could not be scheduled; check the Error Log."),
		}, error_log)

	schedule.normalize_weights(doc.schedule_activities)
	warnings = schedule.validate_schedule(doc.schedule_activities)

	doc.save()
	frappe.db.commit()

	critical = sum(1 for r in doc.schedule_activities if r.is_critical)
	message = _("Schedule generated: {0} activities across {1} phases, {2} on the critical path.").format(
		len(doc.schedule_activities), len(phases), critical
	)
	if info.get("cyclic"):
		message += " " + _("{0} activity(ies) form a dependency loop.").format(len(info["cyclic"]))
	if info.get("dropped_links"):
		message += " " + _("{0} link(s) referenced unknown activities and were ignored.").format(
			len(info["dropped_links"])
		)
	if warnings:
		message += " " + _("{0} consistency warning(s) — use Check Schedule to review.").format(
			len(warnings)
		)

	return {
		"status": "Generated",
		"activities_count": len(doc.schedule_activities),
		"critical_count": critical,
		"warnings_count": len(warnings),
		"message": message,
	}


@frappe.whitelist(methods=["POST"])
def recalculate_schedule(tender_workspace_name):
	"""
	Re-run the CPM pass over the existing activities. No AI involved.

	Use after editing a duration or a predecessor link by hand: dates, float and
	the critical path all follow from the network, so they must be recomputed
	rather than edited.
	"""
	doc = _get_doc(tender_workspace_name)
	if not doc.schedule_activities:
		return {"status": "Empty", "message": _("There is no schedule to recalculate.")}

	try:
		info = schedule.apply_cpm(doc.schedule_activities, _schedule_start_date(doc))
	except schedule.cpm.ScheduleError:
		error_log = _log_tender_error("Tender schedule: CPM could not schedule the network", doc.name)
		return _result_with_error_log(
			{"status": "Failed", "message": _("The network could not be scheduled; check the Error Log.")},
			error_log,
		)

	schedule.normalize_weights(doc.schedule_activities)
	doc.save()
	frappe.db.commit()

	critical = sum(1 for r in doc.schedule_activities if r.is_critical)
	message = _("Recalculated: {0} activities, {1} on the critical path, finishing {2}.").format(
		len(doc.schedule_activities), critical, frappe.utils.formatdate(info.get("project_finish"))
	)
	if info.get("cyclic"):
		message += " " + _("{0} activity(ies) form a dependency loop.").format(len(info["cyclic"]))

	return {
		"status": "Recalculated",
		"critical_count": critical,
		"message": message,
	}


@frappe.whitelist()
def check_schedule(tender_workspace_name):
	"""Re-run the consistency checks and return the warnings for display."""
	doc = _get_doc(tender_workspace_name)
	rows = doc.schedule_activities or []
	if not rows:
		return {"status": "Empty", "warnings": [], "message": _("No schedule activities to check.")}

	warnings = schedule.validate_schedule(rows)
	return {
		"status": "Warnings" if warnings else "Clean",
		"warnings": warnings,
		"activities_count": len(rows),
		"message": (
			_("{0} consistency warning(s) found.").format(len(warnings))
			if warnings
			else _("Schedule is internally consistent.")
		),
	}


# ---------------------------------------------------------------------------
# 7b. Organization structure
# ---------------------------------------------------------------------------
# Tiers are the rows of the printed chart. Keeping them fixed means the chart
# has a predictable shape whatever the project, and the model only has to place
# each role rather than invent a layout.
_ORG_TIERS = (
	"0 = client/contract management (one box, the top of the chart)",
	"1 = project manager (exactly one box)",
	"2 = support functions reporting to the project manager: HSE, QA/QC, "
	"planning and documents, client liaison, procurement and logistics",
	"3 = delivery teams (3-5 boxes), each covering a family of field work",
	"4 = the individual roles inside each tier-3 team",
	"5 = shared labour and logistics pools serving all teams",
)

# A chart box on A4 is roughly 30mm wide; much past this and the title stacks up
# one fragment per line.
_ORG_TITLE_MAX = 45
_ORG_ROLE_TARGET = "20-28"

_ORG_RULES = (
	"RULES:\n"
	"- Every role needs a unique role_code like ORG-010, ORG-020.\n"
	"- reports_to must be the role_code of a box on the tier immediately above. "
	"Tier 0 has an empty reports_to.\n"
	"- box_style follows the tier: 0 Executive, 1 Primary, 2 Support, 3 Team, "
	"4 Role, 5 Labour.\n"
	"- role_title, location, experience, responsibilities and escalation_scope "
	"must be written in ARABIC. Everything else is a code or a number.\n"
	"- headcount is a whole number. Use 0 only for roles called off as needed.\n"
	"- Any role the tender document explicitly requires MUST appear, with "
	"mandated_by_tender = true and a headcount that satisfies the stated "
	"minimum. Do not round those numbers down.\n"
	"- Set escalation_level, escalation_scope and escalation_response on exactly "
	"three roles: the site supervision (Level 1), the project manager (Level 2), "
	"and contract management (Level 3). Response times run same-day, 24 hours, "
	"48 hours.\n"
	"- Base the delivery teams on the work actually in the schedule, not on a "
	"generic construction org chart.\n"
	"- COVER EVERY REGION. If the tender names several cities or zones with a "
	"team count each, every one of them needs field roles with a headcount that "
	"matches its own share. Staffing only the largest city is a rejected bid.\n"
	"- Make the arithmetic add up. If the tender states a team count and a "
	"minimum crew per team, total field headcount must be at least "
	"teams x crew size across all regions. State the resulting number honestly "
	"rather than a smaller round one.\n"
	f"- Keep role_title under {_ORG_TITLE_MAX} characters so it fits a chart box. "
	"Put the detail in responsibilities instead.\n"
	f"- Keep the whole structure to about {_ORG_ROLE_TARGET} roles. Group a "
	"repeated crew into one role with a headcount rather than one role per city "
	"per work type."
)

_ORG_ROW_SCHEMA = (
	'{"role_code": "ORG-010", "role_title": "Arabic job title", "tier": 0, '
	'"reports_to": "", "box_style": "Executive", "headcount": 1, '
	'"location": "Arabic", "experience": "Arabic", '
	'"responsibilities": "Arabic, one sentence", "mandated_by_tender": false, '
	'"escalation_level": "", "escalation_scope": "", "escalation_response": ""}'
)

_ORG_BOX_BY_TIER = {0: "Executive", 1: "Primary", 2: "Support", 3: "Team", 4: "Role", 5: "Labour"}
_ORG_ESCALATION = {"level 1": "Level 1", "level 2": "Level 2", "level 3": "Level 3"}


def _org_generation_context(doc):
	"""Tender context plus the roles the schedule actually uses."""
	context = [_proposal_generation_context(
		doc,
		query="organization manpower staffing roles headcount regions site engineer crew HSE QA QC requirements",
	)]

	counts = {}
	for row in doc.schedule_activities or []:
		resource = (getattr(row, "primary_resource", "") or "").strip()
		if resource:
			counts[resource] = counts.get(resource, 0) + 1
	if counts:
		context.append(
			"\nROLES ALREADY USED BY THE BASELINE SCHEDULE "
			"(role — number of activities). Consolidate spelling variants of the "
			"same job into one role:"
		)
		for resource, n in sorted(counts.items(), key=lambda kv: -kv[1])[:60]:
			context.append(f"- {resource} — {n}")

	phases = orgchart._phase_groups(doc)
	if phases:
		context.append("\nSCHEDULE PHASES:")
		for _stem, activities in phases:
			context.append(f"- {orgchart._phase_title(activities)} ({len(activities)} activities)")

	return "\n".join(context)[:_MAX_DOC_CHARS]


def _normalize_org_row(item):
	"""Coerce one AI role into a child-table row, or None when unusable."""
	code = str(item.get("role_code") or "").strip().upper()
	title = str(item.get("role_title") or "").strip()
	if not code or not title:
		return None

	tier = cint(item.get("tier"))
	tier = min(max(tier, 0), 5)
	style = str(item.get("box_style") or "").strip()
	if style not in orgchart.BOX_COLORS:
		style = _ORG_BOX_BY_TIER[tier]

	escalation = _ORG_ESCALATION.get(str(item.get("escalation_level") or "").strip().lower(), "")

	return {
		"role_code": code[:140],
		"role_title": title[:200],
		"tier": tier,
		"reports_to": str(item.get("reports_to") or "").strip().upper()[:140],
		"box_style": style,
		"headcount": max(cint(item.get("headcount")), 0),
		"location": str(item.get("location") or "").strip()[:100],
		"experience": str(item.get("experience") or "").strip()[:100],
		"responsibilities": str(item.get("responsibilities") or "").strip(),
		"mandated_by_tender": 1 if item.get("mandated_by_tender") else 0,
		"show_in_chart": 1,
		"show_in_table": 1 if tier >= 1 else 0,
		"escalation_level": escalation,
		"escalation_scope": str(item.get("escalation_scope") or "").strip(),
		"escalation_response": str(item.get("escalation_response") or "").strip()[:100],
	}


def _repair_org_links(rows):
	"""
	Drop links that point nowhere and re-parent orphans onto the tier above.

	A broken reports_to would silently detach a whole branch of the chart, so
	the tree is closed here rather than printed with a floating box.
	"""
	by_code = {r["role_code"]: r for r in rows}
	by_tier = {}
	for row in rows:
		by_tier.setdefault(row["tier"], []).append(row)

	repaired = 0
	for row in rows:
		if row["tier"] == 0:
			row["reports_to"] = ""
			continue
		parent = by_code.get(row["reports_to"])
		if parent and parent["tier"] < row["tier"]:
			continue
		# Fall back to the first box on the nearest tier above.
		for tier in range(row["tier"] - 1, -1, -1):
			if by_tier.get(tier):
				row["reports_to"] = by_tier[tier][0]["role_code"]
				repaired += 1
				break
	return repaired


def _ai_org_roles(doc, context):
	"""Ask for the full role list in one call — it is a short, structured answer."""
	system = (
		"You are a proposal manager preparing the project organization structure "
		"for a Saudi government/utility tender. Respond ONLY with valid JSON."
	)
	prompt = (
		"Design the project organization structure for this tender.\n\n"
		"Return a JSON array of roles, each shaped exactly like:\n"
		f"{_ORG_ROW_SCHEMA}\n\n"
		"TIERS:\n" + "\n".join(f"- {t}" for t in _ORG_TIERS) + "\n\n"
		f"{_ORG_RULES}\n\n"
		f"TENDER CONTEXT:\n{context}"
	)
	data = ai_service.complete_json(prompt, system=system, max_tokens=8000, task="extraction")
	items = _as_object_list(data)
	if not items:
		_log_tender_error(
			"Tender AI: no organization roles in reply", doc.name,
			message=f"Sanitized payload:\n{_sanitize_log_text(data)}",
		)
	return items


def _ai_org_narrative(doc, context, roles):
	"""The purpose, basis and closing notes that frame the chart."""
	system = (
		"You are a proposal manager writing the organization structure section of "
		"a Saudi tender submission in Arabic. Respond ONLY with valid JSON."
	)
	titles = "، ".join(r["role_title"] for r in roles[:20])
	prompt = (
		'Return JSON: {"purpose": "...", "basis": ["...", "..."], "notes": ["...", "..."]}\n\n'
		"- purpose: one Arabic paragraph on why this structure was prepared and "
		"how responsibility is divided.\n"
		"- basis: 3-4 Arabic bullets. At least one MUST quote the manpower "
		"minimums the tender document imposes, so the commitment is traceable.\n"
		"- notes: 3-4 Arabic bullets on CVs, subcontractor approval, and scaling "
		"the field teams.\n\n"
		f"ROLES IN THE STRUCTURE: {titles}\n\n"
		f"TENDER CONTEXT:\n{context}"
	)
	data = ai_service.complete_json(prompt, system=system, max_tokens=2000, task="extraction")
	if isinstance(data, dict):
		return data
	_log_tender_error(
		"Tender AI: organization narrative missing",
		doc.name,
		message=f"Payload type: {type(data).__name__}",
	)
	return {}


@frappe.whitelist()
def generate_org_structure(tender_workspace_name):
	"""Queue (re)generation of the project organization structure."""
	doc = _get_doc(tender_workspace_name)

	if not ai_service.is_enabled():
		return _failure_result(
			"AI Not Configured",
			_("AI Settings API key is not configured, so no structure can be generated."),
			"Tender Organization: AI not configured",
			doc.name,
		)

	started = _enqueue_job(
		"organization",
		doc.name,
		"ai_power_tender_management.api.tender_workspace._org_pipeline",
	)
	if not started:
		return {
			"status": "Processing",
			"background": True,
			"message": _("The organization structure is already being generated."),
		}

	return {
		"status": "Processing",
		"background": True,
		"message": _(
			"Generating the organization structure in the background. The form "
			"will update automatically when it finishes."
		),
	}


def _org_pipeline(tender_workspace_name):
	"""Background: build the organization structure."""
	try:
		_publish(tender_workspace_name, _("Reading tender context…"), 10, job_key="organization")
		result = _generate_org_structure_sync(tender_workspace_name)
		state = _job_state_for_result(result)
		error_log = result.get("error_log") or (
			_latest_tender_error(tender_workspace_name) if state in ("failed", "warning") else None
		)
		_publish(
			tender_workspace_name,
			result.get("message") or _("Organization structure generated."),
			100,
			reload=True,
			job_key="organization",
			state=state,
			error_log=error_log,
		)
	except Exception:
		frappe.db.rollback()
		log = _log_tender_error("Tender organization structure generation failed", tender_workspace_name)
		_publish(
			tender_workspace_name,
			_("Organization structure generation failed."),
			100,
			reload=True,
			job_key="organization",
			state="failed",
			error_log=log,
		)


def _generate_org_structure_sync(tender_workspace_name):
	"""(Re)build `organization_roles` and the surrounding narrative."""
	doc = _get_doc(tender_workspace_name)
	context = _org_generation_context(doc)

	_publish(tender_workspace_name, _("Designing the structure…"), 35, job_key="organization")
	items = _ai_org_roles(doc, context)
	if not items:
		error_log = _latest_tender_error(doc.name) or _log_tender_error(
			"Tender Organization: no roles generated",
			doc.name,
			message=f"Context length: {len(context or '')}",
		)
		return _result_with_error_log({
			"status": "AI Failed",
			"message": _("The AI returned no roles; check the Error Log."),
		}, error_log)

	rows, seen = [], set()
	for item in items:
		row = _normalize_org_row(item)
		if not row or row["role_code"] in seen:
			continue
		seen.add(row["role_code"])
		rows.append(row)

	if not rows:
		error_log = _log_tender_error(
			"Tender Organization: no usable roles in AI reply",
			doc.name,
			message=f"Input role count: {len(items)}\nSanitized payload:\n{_sanitize_log_text(items)}",
		)
		return _result_with_error_log(
			{"status": "AI Failed", "message": _("No usable roles in the AI reply; check the Error Log.")},
			error_log,
		)

	repaired = _repair_org_links(rows)

	_publish(tender_workspace_name, _("Writing the narrative…"), 75, job_key="organization")
	narrative = _ai_org_narrative(doc, context, rows)

	doc.set("organization_roles", [])
	for row in rows:
		doc.append("organization_roles", row)

	if narrative.get("purpose"):
		doc.org_purpose = str(narrative["purpose"]).strip()
	for field, key in (("org_basis", "basis"), ("org_notes", "notes")):
		value = narrative.get(key)
		if isinstance(value, list) and value:
			setattr(doc, field, "\n".join(str(v).strip() for v in value if str(v).strip()))

	doc.save()
	frappe.db.commit()

	mandated = sum(1 for r in rows if r["mandated_by_tender"])
	people = sum(r["headcount"] for r in rows)
	message = _("Structure generated: {0} roles, {1} people, {2} required by the tender.").format(
		len(rows), people, mandated
	)
	if repaired:
		message += " " + _("{0} reporting link(s) were repaired.").format(repaired)
	return {"status": "Generated", "roles_count": len(rows), "message": message}


# ---------------------------------------------------------------------------
# 8. Export technical proposal (PLACEHOLDER)
# ---------------------------------------------------------------------------
@frappe.whitelist()
def export_technical_proposal(tender_workspace_name):
	"""
	Generate the technical proposal as a self-contained, print-ready RTL HTML
	document built from `proposal_sections`, attach it to the tender, and return
	its file_url.

	HTML (not PDF) is used because it renders Arabic reliably in any browser and
	needs no server-side PDF engine; the user can Print → Save as PDF.

	TODO(export): when a PDF engine (wkhtmltopdf / weasyprint) is available,
	convert this same HTML via `frappe.utils.pdf.get_pdf` and save a .pdf instead.
	"""
	doc = _get_doc(tender_workspace_name)

	if not doc.proposal_sections:
		frappe.throw(_("Please Generate Proposal Sections first."))

	try:
		html = _build_technical_html(doc)
		filename = f"Technical-Proposal-{doc.name}.html"
		file_url = _save_export(filename, html, doc.name)
	except Exception:
		frappe.db.rollback()
		error_log = _log_tender_error("Tender Export: technical proposal failed", doc.name)
		return _result_with_error_log({
			"status": "Failed",
			"sections_count": len(doc.proposal_sections),
			"message": _("Technical proposal export failed. Open the Error Log below, then retry."),
		}, error_log)

	return {
		"status": "Ready",
		"sections_count": len(doc.proposal_sections),
		"file_url": file_url,
		"message": _("Technical proposal generated ({0} sections).").format(len(doc.proposal_sections)),
	}


# ---------------------------------------------------------------------------
# 9. Export financial proposal (PLACEHOLDER)
# ---------------------------------------------------------------------------
@frappe.whitelist()
def export_financial_proposal(tender_workspace_name):
	"""
	Generate the financial proposal as a real XLSX priced-BOQ workbook built
	from `boq_items` (via openpyxl), attach it to the tender, and return its
	file_url. Includes a grand-total row.
	"""
	doc = _get_doc(tender_workspace_name)

	if not doc.boq_items:
		frappe.throw(_("Please Extract BOQ first."))

	try:
		content, grand_total = _build_financial_xlsx(doc)
		filename = f"Financial-Proposal-{doc.name}.xlsx"
		file_url = _save_export(filename, content, doc.name)
	except Exception:
		frappe.db.rollback()
		error_log = _log_tender_error("Tender Export: financial proposal failed", doc.name)
		return _result_with_error_log({
			"status": "Failed",
			"items_count": len(doc.boq_items),
			"message": _("Financial proposal export failed. Open the Error Log below, then retry."),
		}, error_log)

	return {
		"status": "Ready",
		"items_count": len(doc.boq_items),
		"grand_total": grand_total,
		"file_url": file_url,
		"message": _("Financial proposal generated ({0} items).").format(len(doc.boq_items)),
	}


# ---------------------------------------------------------------------------
# Export helpers
# ---------------------------------------------------------------------------
def _save_export(filename, content, tender_name):
	"""Attach a generated export file to the tender and return its file_url.

	Re-running an export replaces the previous file with the same name so the
	attachment list does not grow on every click.
	"""
	# Frappe may append a random suffix to the stored file_name on disk
	# collisions (e.g. "...-TND-00002ab12cd.xlsx"), so match by prefix and scope
	# strictly to this tender's attachments.
	base = filename.rsplit(".", 1)[0]
	existing = frappe.get_all(
		"File",
		filters={
			"attached_to_doctype": "Tender Workspace",
			"attached_to_name": tender_name,
			"file_name": ["like", f"{base}%"],
		},
		pluck="name",
	)
	for name in existing:
		frappe.delete_doc("File", name, ignore_permissions=True, force=True)

	_file = save_file(filename, content, "Tender Workspace", tender_name, is_private=1)
	return _file.file_url


def _tender_display_name(doc, arabic=False):
	"""Tender name for display; prefers the Arabic name in Arabic output."""
	if arabic or (frappe.local.lang or "").startswith("ar"):
		return doc.tender_name_ar or doc.tender_name or ""
	return doc.tender_name or ""


def _build_technical_html(doc):
	"""Build a styled, print-ready RTL Arabic HTML technical proposal."""
	esc = frappe.utils.escape_html
	display_name = _tender_display_name(doc, arabic=True)

	# The tender name is already the cover subtitle — do not repeat it here.
	meta = [
		(_("Tender No"), doc.tender_number),
		(_("Client"), doc.client_name),
		(_("Closing Date"), frappe.utils.formatdate(doc.closing_date) if doc.closing_date else ""),
	]
	meta_html = "".join(
		f"<div class='meta-row'><span class='meta-label'>{esc(label)}</span>"
		f"<span class='meta-value'>{esc(value or '—')}</span></div>"
		for label, value in meta
	)

	# The timeline section is rendered from the schedule rows rather than its own
	# prose. It falls back to the text when no schedule has been generated yet.
	schedule_html = schedule.build_schedule_html(doc)

	sections_html = ""
	for i, s in enumerate(doc.proposal_sections, start=1):
		if s.section_type == SCHEDULE_SECTION and schedule_html:
			body = schedule_html
			wide = " wide"
		else:
			# `content` comes from a Text Editor field -> already HTML, do not escape.
			body = s.content or ""
			wide = ""
		sections_html += (
			f"<section class='sec{wide}'>"
			f"<h2>{i}. {esc(s.title or s.section_type)}</h2>"
			f"<div class='content'>{body}</div>"
			f"</section>"
		)

	# The schedule hangs off the Project Timeline section. If that row was never
	# generated the schedule would silently vanish from the proposal, so give it
	# a section of its own instead.
	if schedule_html and not any(s.section_type == SCHEDULE_SECTION for s in doc.proposal_sections):
		sections_html += (
			f"<section class='sec wide'>"
			f"<h2>{len(doc.proposal_sections) + 1}. {esc(_proposal_title(SCHEDULE_SECTION))}</h2>"
			f"<div class='content'>{schedule_html}</div>"
			f"</section>"
		)

	# Contents list. Page numbers are deliberately absent: pagination is decided
	# by the print engine, so any number printed here would be a guess.
	toc_items = [esc(s.title or s.section_type) for s in doc.proposal_sections]
	if schedule_html and not any(s.section_type == SCHEDULE_SECTION for s in doc.proposal_sections):
		toc_items.append(esc(_proposal_title(SCHEDULE_SECTION)))
	toc_html = "".join(
		f"<li><span class='n'>{i}</span>{title}</li>" for i, title in enumerate(toc_items, start=1)
	)

	# A Gantt spanning months needs the long edge of the page; only switch the
	# document to landscape when there is actually a chart to fit.
	page_css = "@page { size: A4 landscape; margin: 12mm 12mm 16mm; }" if schedule_html else "@page { margin: 16mm 16mm 20mm; }"

	footer_bits = [b for b in (doc.tender_number, display_name or doc.name) if b]
	footer_text = esc(" · ".join(footer_bits))

	return f"""<!doctype html>
<html dir="rtl" lang="ar">
<head>
<meta charset="utf-8">
<title>{esc(_('Technical Proposal'))} — {esc(display_name or doc.name)}</title>
<style>
	* {{ box-sizing: border-box; }}
	body {{ font-family: 'Tahoma','Segoe UI','Arial',sans-serif; color: #1f2430; margin: 0; padding: 32px 40px; line-height: 1.7; }}
	.cover {{ background: linear-gradient(120deg,#4f46e5,#7c3aed); color:#fff; border-radius: 14px; padding: 28px 30px; margin-bottom: 26px; }}
	.cover h1 {{ margin: 0 0 6px; font-size: 26px; }}
	.cover .subtitle {{ opacity: .85; font-size: 14px; }}
	.meta {{ margin-top: 18px; display: grid; grid-template-columns: 1fr 1fr; gap: 6px 24px; }}
	.meta-row {{ display: flex; gap: 8px; font-size: 13px; }}
	.meta-label {{ opacity: .8; min-width: 90px; }}
	.meta-value {{ font-weight: 600; }}
	.sec {{ margin-bottom: 22px; padding: 18px 20px; border: 1px solid #e6e8ee; border-radius: 12px; page-break-inside: avoid; }}
	.sec h2 {{ margin: 0 0 10px; font-size: 17px; color: #4f46e5; border-bottom: 2px solid #eef0f6; padding-bottom: 8px; }}
	.content {{ font-size: 14px; }}
	/* The schedule needs the full width and may run over several pages. */
	.sec.wide {{ padding: 14px; page-break-inside: auto; }}
	.sec.wide .content {{ font-size: 12px; }}

	/* Cover occupies its own sheet; the contents list follows on the next. */
	.cover-page {{ min-height: 86vh; display: flex; flex-direction: column; justify-content: center;
		page-break-after: always; }}
	.cover-page .doc-type {{ font-size: 13px; letter-spacing: .18em; opacity: .75; }}
	.cover-page h1 {{ font-size: 30px; margin: 8px 0 18px; }}
	.toc {{ page-break-after: always; }}
	.toc h2 {{ font-size: 18px; color: #4f46e5; margin: 0 0 12px; }}
	.toc ol {{ list-style: none; padding: 0; margin: 0; font-size: 14px; }}
	.toc li {{ padding: 7px 0; border-bottom: 1px dashed #d9dde6; }}
	.toc .n {{ display: inline-block; min-width: 26px; color: #7c3aed; font-weight: 700; }}

	/* Repeats on every printed page in browsers; wkhtmltopdf draws its own. */
	.page-footer {{ position: fixed; bottom: 0; left: 0; right: 0; font-size: 9px; color: #8b93a3;
		border-top: 1px solid #e6e8ee; padding: 4px 6px; background: #fff; }}
	.page-footer .r {{ float: left; }}

	@media print {{
		body {{ padding: 0; }}
		.sec {{ break-inside: avoid; }}
		.sec.wide {{ break-inside: auto; }}
		.cover-page {{ min-height: 0; height: 92vh; }}
	}}
	{page_css}
	{schedule.SCHEDULE_CSS}
</style>
</head>
<body>
	<div class="page-footer">{footer_text}<span class="r">{esc(_('Technical Proposal'))}</span></div>

	<div class="cover-page">
		<div class="cover">
			<div class="doc-type">{esc(_('Technical Proposal'))}</div>
			<h1>{esc(display_name)}</h1>
			<div class="meta">{meta_html}</div>
		</div>
	</div>

	<div class="toc">
		<h2>{esc(_('Contents'))}</h2>
		<ol>{toc_html}</ol>
	</div>

	{sections_html}
</body>
</html>"""


def _build_financial_xlsx(doc):
	"""Build a priced-BOQ XLSX workbook; return (bytes, grand_total)."""
	import openpyxl
	from openpyxl.styles import Alignment, Font, PatternFill

	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "Financial Proposal"

	# Title row
	ws["A1"] = f"Financial Proposal — {_tender_display_name(doc) or doc.name}"
	ws["A1"].font = Font(size=14, bold=True)
	ws.merge_cells("A1:J1")

	headers = [
		"Line Type", "Item No", "Parent Item No", "Description", "Description (English)",
		"Unit", "Quantity", "Unit Price", "Total", "Specification",
	]
	header_row = 3
	ws.append([])  # row 2 spacer
	ws.append(headers)
	header_fill = PatternFill("solid", fgColor="4F46E5")
	for col in range(1, len(headers) + 1):
		cell = ws.cell(row=header_row, column=col)
		cell.font = Font(color="FFFFFF", bold=True)
		cell.fill = header_fill
		cell.alignment = Alignment(horizontal="center")

	subtotal = 0.0
	for row in doc.boq_items:
		is_heading = (row.line_type or "Item") == "Section Heading"
		total = 0 if is_heading else flt(row.quantity) * flt(row.unit_price)
		subtotal += total
		ws.append([
			row.line_type or "Item",
			row.item_no,
			row.parent_item_no,
			row.description,
			row.description_en,
			row.unit,
			0 if is_heading else flt(row.quantity),
			0 if is_heading else flt(row.unit_price),
			total,
			row.specification,
		])
		# Right-align + RTL reading order for the Arabic text columns
		# (Description = D, Description EN = E, Specification = J).
		rtl = Alignment(horizontal="right", readingOrder=2, vertical="top", wrap_text=True)
		for col in (4, 5, 10):
			ws.cell(row=ws.max_row, column=col).alignment = rtl

	vat_rate = flt(doc.vat_rate if doc.vat_rate is not None else 15)
	vat_amount = subtotal * vat_rate / 100
	grand_total = subtotal + vat_amount

	for label, amount in (
		("Subtotal", subtotal),
		(f"VAT ({vat_rate:g}%)", vat_amount),
		("Grand Total", grand_total),
	):
		ws.append(["", "", "", "", "", "", "", label, amount, ""])
	last = ws.max_row
	for row_idx in range(last - 2, last + 1):
		ws.cell(row=row_idx, column=8).font = Font(bold=True)
		ws.cell(row=row_idx, column=9).font = Font(bold=True)

	# Column widths
	for col, width in zip("ABCDEFGHIJ", [14, 12, 14, 42, 42, 10, 12, 14, 14, 30]):
		ws.column_dimensions[col].width = width

	buf = io.BytesIO()
	wb.save(buf)
	return buf.getvalue(), grand_total


# ---------------------------------------------------------------------------
# 6. Processing summary
# ---------------------------------------------------------------------------
@frappe.whitelist()
def get_processing_summary(tender_workspace_name):
	"""Return a compact status/counts summary for the processing result card."""
	doc = _get_doc(tender_workspace_name)

	tender_doc = _find_document(doc, TENDER_DOC_TYPES)
	boq_doc = _find_document(doc, BOQ_DOC_TYPES)

	dangerous = len([r for r in doc.ai_summary if r.summary_type == "Dangerous Clause"])
	missing = len([r for r in doc.ai_summary if r.summary_type == "Missing Information"])

	return {
		"tender_status": doc.status,
		"tender_document_status": tender_doc.ai_status if tender_doc else "Not Uploaded",
		"boq_status": boq_doc.ai_status if boq_doc else "Not Uploaded",
		"ai_summary_count": len(doc.ai_summary),
		"dangerous_clauses_count": dangerous,
		"missing_information_count": missing,
		"boq_items_count": len(doc.boq_items),
		"proposal_sections_count": len(doc.proposal_sections),
		"tender_summary_created": any(r.summary_type == "Tender Summary" for r in doc.ai_summary),
		"ai_enabled": ai_service.is_enabled(),
		# Every AI step runs in a background job, so "busy" must be decided from
		# the job list, not from the per-document ai_status flags.
		"jobs": get_background_jobs(doc.name)["jobs"],
	}


# ---------------------------------------------------------------------------
# LLM helpers — build prompts, call the model (via AI Settings), parse output.
# Each returns None on any problem so callers fall back to placeholders.
# ---------------------------------------------------------------------------
# Truncate very large documents before sending to the LLM (context safety).
_MAX_DOC_CHARS = 12000
_SUMMARY_TYPES = list(SUMMARY_BLUEPRINT.keys())


def _ai_summary_rows(doc, document_text):
	"""Ask the LLM to extract structured tender-summary rows from the text."""
	text = (document_text or "").strip()
	if len(text) < 100:
		return None

	system = (
		"You are a tender analyst. Extract structured insights from tender "
		"documents. Respond ONLY with valid JSON, no prose."
	)
	prompt = (
		"From the tender document below, extract insights as a JSON array. "
		"Each element: {\"summary_type\": one of "
		f"{_SUMMARY_TYPES}, \"extracted_text\": short finding, "
		"\"page_number\": string or empty}. Include several Dangerous Clause and "
		"Missing Information items when present.\n\n"
		f"DOCUMENT:\n{text[:_MAX_DOC_CHARS]}"
	)
	data = ai_service.complete_json(prompt, system=system, task="extraction")
	if not isinstance(data, list):
		return None

	rows = []
	for item in data:
		if not isinstance(item, dict):
			continue
		stype = item.get("summary_type")
		if stype in _SUMMARY_TYPES and item.get("extracted_text"):
			rows.append(item)
	return rows or None


# BOQ extraction tuning.
#   - Below this overall confidence (or on any deterministic price-provenance
#     issue) a second "verifier" pass is run on the same model. Above it, the
#     draft is trusted and the second call is skipped to conserve the API key.
_BOQ_VERIFY_THRESHOLD = 75.0
# Arabic-Indic digits -> ASCII, so numeric provenance checks work on Arabic docs.
_ARABIC_DIGITS = str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789")


def _normalize_boq_row(item):
	"""Coerce one raw LLM item dict into the canonical boq_items row shape."""
	qty = flt(item.get("quantity"))
	price = flt(item.get("unit_price"))
	line_type = str(item.get("line_type") or "Item")
	if line_type not in ("Item", "Section Heading"):
		line_type = "Item"
	total = qty * price if line_type == "Item" else 0
	return {
		"line_type": line_type,
		"item_no": str(item.get("item_no") or ""),
		"parent_item_no": str(item.get("parent_item_no") or ""),
		"description": str(item.get("description") or ""),
		"description_en": str(item.get("description_en") or ""),
		"unit": str(item.get("unit") or ""),
		"quantity": qty,
		"unit_price": price,
		"total": total,
		"specification": str(item.get("specification") or ""),
		"source_page": str(item.get("source_page") or ""),
		"extraction_confidence": flt(item.get("extraction_confidence")),
	}


def _number_in_text(value, normalized_src):
	"""True if a numeric value plausibly appears in the (normalized) source text."""
	v = flt(value)
	if not v:
		return True
	forms = {("%g" % v), ("%.2f" % v)}
	if v == int(v):
		forms.add(str(int(v)))
	# Compare with separators stripped so "1,500.00" matches "1500".
	return any(f and f.replace(".", "") in normalized_src for f in forms)


def _apply_price_provenance(rows, source_text):
	"""Deterministic guard (no API): drop any unit_price not found in the source.

	This is the cheap safety net that kills fabricated prices even if the model
	invents them — an unpriced BOQ must never come back with prices.
	"""
	normalized_src = (source_text or "").translate(_ARABIC_DIGITS).replace(",", "").replace(" ", "").replace(".", "")
	issues = []
	for r in rows:
		price = flt(r.get("unit_price"))
		if price and not _number_in_text(price, normalized_src):
			issues.append(
				f"Removed price {price} for '{(r.get('description') or '')[:30]}' — not present in source"
			)
			r["unit_price"] = 0
			r["total"] = 0
	return issues


def _ai_boq_rows(boq_text):
	"""Pass 1 (Extractor): ask the LLM to extract BOQ line items from raw text."""
	text = (boq_text or "").strip()
	if len(text) < 50:
		return None

	system = "You extract Bill of Quantities line items. Respond ONLY with valid JSON."
	prompt = (
		"Extract the BOQ line items from the SOURCE below as a JSON array. Keep the "
		"output compact — one element per line item, no extra prose. Each element: "
		"{\"line_type\": \"Item\" or \"Section Heading\", \"item_no\": str, "
		"\"parent_item_no\": str, \"description\": str (keep the ORIGINAL language, "
		"do NOT translate), \"unit\": str, \"quantity\": number, \"unit_price\": number, "
		"\"source_page\": str, \"extraction_confidence\": 0-100}. "
		"The source may contain title rows, merged cells, shifted columns and "
		"bilingual headers — infer the table structure yourself; do not assume a "
		"fixed layout. CRITICAL: never invent prices or quantities. If a unit price "
		"is blank/absent in the source, set unit_price to 0 (many BOQs are unpriced). "
		"For headings or subtotal/VAT lines set line_type to \"Section Heading\" and "
		"quantity/unit_price to 0.\n\n"
		f"SOURCE:\n{text[:_MAX_DOC_CHARS]}"
	)
	data = ai_service.complete_json(prompt, system=system, max_tokens=8000, task="extraction")
	if not isinstance(data, list):
		return None

	rows = [_normalize_boq_row(item) for item in data if isinstance(item, dict) and item.get("description")]
	return rows or None


def _ai_boq_verify(source_text, draft_rows):
	"""Pass 2 (Verifier): audit the draft against the source on the same model.

	Runs only when Pass 1 is low-confidence or the provenance guard flagged
	something. Returns {"rows", "confidence", "issues"} or None on failure.
	"""
	text = (source_text or "").strip()
	if not text or not draft_rows:
		return None

	system = "You audit an extracted Bill of Quantities against its source. Respond ONLY with valid JSON."
	prompt = (
		"You are given the SOURCE text of a BOQ and a DRAFT extraction of its line "
		"items. Audit the draft against the source and return corrected data as a "
		"JSON object: {\"rows\": [ ...same element schema as the draft... ], "
		"\"confidence\": 0-100, \"issues\": [str]}. Rules: (1) Remove draft rows that "
		"do not exist in the source. (2) Add line items that are in the source but "
		"missing from the draft. (3) NEVER invent prices or quantities — if a unit "
		"price is not explicitly in the source, set unit_price to 0. (4) Keep "
		"descriptions in their original language. Set confidence to how well the "
		"corrected rows match the source.\n\n"
		f"SOURCE:\n{text[:_MAX_DOC_CHARS]}\n\n"
		f"DRAFT:\n{frappe.as_json(draft_rows)[:_MAX_DOC_CHARS]}"
	)
	data = ai_service.complete_json(prompt, system=system, max_tokens=8000, task="validation")
	if not isinstance(data, dict) or not isinstance(data.get("rows"), list):
		return None

	rows = [_normalize_boq_row(item) for item in data["rows"] if isinstance(item, dict) and item.get("description")]
	return {"rows": rows, "confidence": flt(data.get("confidence")), "issues": data.get("issues") or []}


def _extract_boq_from_text(source_text, source_label="", allow_verify=True):
	"""Format-agnostic BOQ extraction orchestrator (single model, up to 2 passes).

	Pass 1 extracts, a free deterministic guard strips fabricated prices, and a
	conditional Pass 2 verifies only when confidence is low or a price was
	flagged. Returns (rows, confidence, issues).
	"""
	text = (source_text or "").strip()
	if len(text) < 30:
		return [], 0.0, ["Source text too short to extract."]

	rows = _ai_boq_rows(text) or []
	if not rows:
		return [], 0.0, ["Extractor returned no rows."]

	issues = _apply_price_provenance(rows, text)

	confs = [flt(r.get("extraction_confidence")) for r in rows if flt(r.get("extraction_confidence"))]
	confidence = (sum(confs) / len(confs)) if confs else 60.0
	if issues:
		confidence = min(confidence, 55.0)

	# Pass 2 only when it can actually help (low confidence or a flagged price).
	if allow_verify and (confidence < _BOQ_VERIFY_THRESHOLD or issues):
		verified = _ai_boq_verify(text, rows)
		if verified and verified.get("rows"):
			rows = verified["rows"]
			issues += _apply_price_provenance(rows, text)  # re-guard corrected rows
			if verified.get("confidence"):
				confidence = flt(verified["confidence"])
			if verified.get("issues"):
				issues += [str(i) for i in verified["issues"]]

	return rows, confidence, issues


def _filter_boq_rows(rows):
	"""Reduce to real line items for an items-only BOQ.

	Keeps every ``Item`` row plus any ``Section Heading`` that is referenced as
	a parent by some item (so ``parent_item_no`` links never orphan). Drops the
	standalone headings — document title, subtotal/VAT/total footer rows — that
	nothing references and that the app recomputes on its own.
	"""
	referenced_parents = {
		(r.get("parent_item_no") or "").strip()
		for r in rows
		if (r.get("parent_item_no") or "").strip()
	}
	kept = []
	for r in rows:
		if (r.get("line_type") or "Item") != "Section Heading":
			kept.append(r)
			continue
		item_no = (r.get("item_no") or "").strip()
		if item_no and item_no in referenced_parents:
			kept.append(r)
	return kept


def _rows_from_excel_deterministic(file_url):
	"""Free, no-AI spreadsheet parse (used only when AI is not configured).

	Best-effort column mapping via document_parser; rows whose columns could
	not be detected are skipped rather than fabricated.
	"""
	rows = []
	for r in document_parser.extract_rows_from_excel(file_url):
		if "raw" in r:
			continue
		rows.append(_normalize_boq_row(r))
	return rows


def _proposal_generation_context(doc, query=None):
	"""Build compact tender context for detailed proposal section generation."""
	# getattr: optional/newer fields may be missing on a site that has not run
	# `bench migrate` yet — a missing name must not abort proposal generation.
	context_lines = [
		f"Tender: {doc.tender_name or ''}",
		f"Tender (Arabic): {getattr(doc, 'tender_name_ar', None) or ''}",
		f"Tender No: {doc.tender_number or ''}",
		f"Client: {doc.client_name or ''}",
	]

	summary_rows = [r for r in getattr(doc, "ai_summary", []) if getattr(r, "extracted_text", None)]
	if summary_rows:
		context_lines.append("\nAI SUMMARY:")
	for r in summary_rows[:30]:
		page = f" (page {r.page_number})" if getattr(r, "page_number", None) else ""
		context_lines.append(f"- {r.summary_type}{page}: {r.extracted_text}")

	boq_rows = [r for r in getattr(doc, "boq_items", []) if getattr(r, "description", None)]
	if boq_rows:
		context_lines.append("\nBOQ ITEMS:")
	for r in boq_rows[:30]:
		line_type = getattr(r, "line_type", None) or "Item"
		parent = f", parent {r.parent_item_no}" if getattr(r, "parent_item_no", None) else ""
		quantity = f", qty {r.quantity}" if getattr(r, "quantity", None) else ""
		unit = f" {r.unit}" if getattr(r, "unit", None) else ""
		total = f", total {r.total}" if getattr(r, "total", None) else ""
		spec = f", spec: {r.specification}" if getattr(r, "specification", None) else ""
		context_lines.append(f"- [{line_type}] {r.item_no or ''}{parent} {r.description}{quantity}{unit}{total}{spec}")

	retrieval_query = query or (
		"scope methodology implementation equipment organization QA QC HSE compliance risk "
		"submission commercial technical requirements BOQ quantities schedule manpower"
	)
	retrieved = _retrieved_knowledge_context(doc, retrieval_query)
	if retrieved:
		context_lines.append(retrieved)

	return "\n".join(context_lines)[:_MAX_DOC_CHARS]


def _flatten_section_payload(data):
	"""
	Reduce the model's JSON reply to a flat {label: html} map.

	The reply is *usually* {"Section Name": "<p>…"} but the model also returns
	{"section": …, "content": …} pairs, a list of those, or the same nested one
	level under "sections". Walking the structure covers all of them.
	"""
	found = {}

	def _add(label, body):
		if isinstance(body, str) and body.strip():
			found.setdefault(str(label).strip(), body.strip())

	def _walk(node):
		if isinstance(node, dict):
			# A {"section": ..., "content": ...} record describes one section.
			label = (
				node.get("section") or node.get("section_type")
				or node.get("title") or node.get("name")
			)
			body = (
				node.get("content") or node.get("html")
				or node.get("body") or node.get("text")
			)
			if isinstance(label, str) and isinstance(body, str):
				_add(label, body)
				return
			for key, value in node.items():
				if isinstance(value, str):
					_add(key, value)
				else:
					_walk(value)
		elif isinstance(node, list):
			for item in node:
				_walk(item)

	_walk(data)
	return found


def _normalise_section_key(value):
	"""Fold case, spacing and punctuation so near-miss keys still match."""
	return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


# Below this, a lone string in the reply is too short to be real section prose.
_MIN_SECTION_CHARS = 80


def _match_section_content(flat, section, single=False):
	"""Find one section's content in the flattened reply, tolerating key drift."""
	if not flat:
		return None

	if section in flat:
		return flat[section]

	normalised = {_normalise_section_key(k): v for k, v in flat.items()}
	hit = normalised.get(_normalise_section_key(section))
	if hit:
		return hit

	# The model sometimes keys the content by the Arabic title it just wrote.
	arabic = PROPOSAL_SECTION_TITLES_AR.get(section)
	if arabic and arabic in flat:
		return flat[arabic]

	# Only one section was requested and the reply carries exactly one body of
	# real prose — whatever it called it, that is the answer.
	if single:
		bodies = [v for v in flat.values() if len(v) >= _MIN_SECTION_CHARS]
		if len(bodies) == 1:
			return bodies[0]

	return None


def _ai_proposal_map(doc, sections=None, on_progress=None):
	"""Ask the LLM for detailed Arabic proposal content keyed by section name.

	`on_progress(done, total)` — optional callback invoked after each batch so a
	background job can report how far along it is.
	"""
	sections = [
		_normalize_proposal_section(section)
		for section in (sections or PROPOSAL_SECTIONS)
		if _normalize_proposal_section(section) in PROPOSAL_SECTIONS
	]
	if not sections:
		return None
	context = _proposal_generation_context(
		doc,
		query=" ".join(sections) + " " + " ".join(PROPOSAL_SECTION_GUIDANCE.get(section, "") for section in sections),
	)

	system = (
		"You are a senior technical proposal writer for government and utility "
		"tenders in Saudi Arabia. Write formal Arabic. Respond ONLY with valid JSON."
	)

	result = {}
	for start in range(0, len(sections), _PROPOSAL_BATCH_SIZE):
		batch = sections[start : start + _PROPOSAL_BATCH_SIZE]
		section_list = "\n".join(f"- {section}" for section in batch)
		guidance = "\n".join(
			f"- {section}: {PROPOSAL_SECTION_GUIDANCE.get(section, '')}"
			for section in batch
		)
		prompt = (
			"Write detailed Arabic technical proposal content for EACH section "
			f"listed below. Target {_PROPOSAL_SECTION_WORD_TARGET} Arabic words per "
			"section. Use tender-specific details from the context where available; "
			"where the context is missing, state practical assumptions without "
			"claiming unsupported facts. Return short HTML fragments only, using "
			"<p>, <ul>, <li>, and <table> where useful. Do not return Markdown, code "
			"fences, or a full HTML document. Return a valid JSON object mapping each "
			"exact English section name to its Arabic HTML content string.\n\n"
			f"SECTIONS:\n{section_list}\n\n"
			f"SECTION GUIDANCE:\n{guidance}\n\n"
			f"TENDER CONTEXT:\n{context}"
		)
		data = ai_service.complete_json(
			prompt,
			system=system,
			max_tokens=_PROPOSAL_BATCH_MAX_TOKENS,
			task="arabic_proposal",
		)
		if isinstance(data, dict):
			flat = _flatten_section_payload(data)
			for section in batch:
				content = _match_section_content(flat, section, single=len(batch) == 1)
				if content:
					result[section] = content
				else:
					# Silently dropping this surfaced as "AI Failed" with nothing in
					# the Error Log, which is undiagnosable. Record what came back.
					_log_tender_error(
						"Tender AI: proposal section not found in reply",
						doc.name,
						message=(
							f"Section: {section}\nKeys returned: {sorted(flat)}\n\n"
							f"Sanitized payload:\n{_sanitize_log_text(data)}"
						),
					)
		else:
			_log_tender_error(
				"Tender AI: proposal batch returned no JSON object",
				doc.name,
				message=f"Sections: {', '.join(batch)}\nPayload type: {type(data).__name__}",
			)

		if on_progress:
			on_progress(min(start + len(batch), len(sections)), len(sections))

	return result or None


# ---------------------------------------------------------------------------
# Vision helpers — for PDFs with no text layer (scanned/image). Send the PDF
# natively to the model (Anthropic) instead of extracted text.
# ---------------------------------------------------------------------------
def _ai_summary_rows_pdf(doc, file_url):
	"""Extract structured tender-summary rows by reading the PDF natively."""
	system = (
		"You are a tender analyst reading a scanned Arabic tender document (كراسة الشروط). "
		"Extract structured insights. Respond ONLY with valid JSON, no prose."
	)
	prompt = (
		"Read the attached tender PDF and extract insights as a JSON array. Each element: "
		"{\"summary_type\": one of "
		f"{_SUMMARY_TYPES}, \"extracted_text\": short finding (Arabic ok), "
		"\"page_number\": string}. Include several Dangerous Clause (penalties, "
		"guarantees) and Missing Information items when present."
	)
	data = ai_service.complete_pdf_json(file_url, prompt, system=system, task="extraction")
	if not isinstance(data, list):
		return None

	rows = []
	for item in data:
		if isinstance(item, dict) and item.get("summary_type") in _SUMMARY_TYPES and item.get("extracted_text"):
			rows.append(item)
	return rows or None


def _ai_boq_rows_pdf(file_url):
	"""Extract BOQ line items by reading the PDF natively (vision)."""
	system = "You extract Bill of Quantities line items from a scanned PDF. Respond ONLY with valid JSON."
	prompt = (
		"Read the attached tender PDF and find the BOQ / quantities table "
		"(جدول الكميات والأسعار). Return the line items as a JSON array. Each element: "
		"{\"line_type\": \"Item\" or \"Section Heading\", \"item_no\": str, "
		"\"parent_item_no\": str, \"description\": str, \"description_en\": English "
		"translation if useful, \"unit\": str, \"quantity\": number, \"unit_price\": number, "
		"\"specification\": str, \"source_page\": str, \"extraction_confidence\": 0-100}. "
		"If prices are absent (unpriced BOQ), set unit_price to 0. For headings or "
		"subtotal/VAT lines, set line_type to \"Section Heading\" and quantity/unit_price to 0."
	)
	data = ai_service.complete_pdf_json(file_url, prompt, system=system, task="extraction")
	if not isinstance(data, list):
		return None

	rows = [_normalize_boq_row(item) for item in data if isinstance(item, dict) and item.get("description")]
	return rows or None


# ---------------------------------------------------------------------------
# OCR pipeline (background) — for scanned PDFs with no text layer.
#   OCR the PDF locally -> chunk the text -> send chunks to the LLM (throttled
#   under the rate limit) -> aggregate structured rows. Runs via frappe.enqueue.
# ---------------------------------------------------------------------------
# Chars per chunk (~3k tokens) — small enough to stay under the per-minute limit.
_CHUNK_CHARS = 12000
# Keywords that mark the BOQ / quantities pages in the OCR text.
_BOQ_MARKERS = ("جدول الكميات", "الكمية", "وصف البند", "BOQ", "FLOW METER", "QUANTITY", "UNIT PRICE")


def _chunk_text(text, max_chars=_CHUNK_CHARS):
	"""Split text into chunks of at most max_chars, breaking on page boundaries."""
	text = text or ""
	if len(text) <= max_chars:
		return [text] if text.strip() else []

	chunks, current = [], ""
	for block in text.split("\n\n"):
		if len(current) + len(block) + 2 > max_chars and current:
			chunks.append(current)
			current = ""
		if len(block) > max_chars:
			for i in range(0, len(block), max_chars):
				chunks.append(block[i : i + max_chars])
			continue
		current += ("\n\n" if current else "") + block
	if current.strip():
		chunks.append(current)
	return chunks


# ---------------------------------------------------------------------------
# Background job tracking
#   Long AI/OCR steps run via frappe.enqueue. Progress is pushed to open forms
#   over realtime AND mirrored into the cache, so a user who reloads (or opens
#   the tender in a second tab) still sees what is running.
# ---------------------------------------------------------------------------
# Job keys -> label shown in the form's "Background Processes" panel. Kept as
# plain strings (translated at render time, not at import time).
BACKGROUND_JOBS = {
	"tender-info": "Tender Info Extraction",
	"analyze": "Document Analysis (OCR + AI)",
	"boq": "BOQ Extraction (OCR + AI)",
	"proposal": "Proposal Sections",
	"proposal-section": "Proposal Section",
	"schedule": "Baseline Schedule",
	"organization": "Organization Structure",
}

# Keep finished/failed entries around long enough to be seen after a reload.
_JOB_CACHE_TTL = 24 * 60 * 60
# A "running" entry that has not been touched for this long, and whose RQ job is
# gone, is treated as dead (worker crash / restart) rather than still running.
_JOB_STALE_SECONDS = 120


def _job_id(job_key, name):
	return f"tender-{job_key}-{name}"


def _job_cache_key(name):
	return f"tender_background_jobs::{name}"


def _read_jobs(name):
	return frappe.cache().get_value(_job_cache_key(name)) or {}


def _set_job_state(name, job_key, state, message, progress, label=None, error_log=None):
	"""Record a background job's state in the cache (survives page reloads)."""
	jobs = _read_jobs(name)
	jobs[job_key] = {
		"key": job_key,
		"label": label or BACKGROUND_JOBS.get(job_key, job_key),
			"state": state,  # queued | running | done | warning | failed
		"message": message,
		"progress": progress,
		"updated": frappe.utils.now(),
		# Name of the Error Log this failure produced, so the panel can link
		# straight to it instead of telling the user to go and find it.
		"error_log": error_log,
	}
	frappe.cache().set_value(_job_cache_key(name), jobs, expires_in_sec=_JOB_CACHE_TTL)


def _publish(name, message, progress, reload=False, job_key="analyze", state=None, error_log=None):
	"""Push OCR/AI progress to any open Tender Workspace form."""
	state = state or ("done" if progress >= 100 else "running")
	_set_job_state(name, job_key, state, message, progress, error_log=error_log)
	frappe.publish_realtime(
		"tender_analyze_progress",
		{
			"name": name,
			"message": message,
			"progress": progress,
			"reload": reload,
			"job_key": job_key,
			"label": BACKGROUND_JOBS.get(job_key, job_key),
			"state": state,
			"error_log": error_log,
		},
		doctype="Tender Workspace",
		docname=name,
	)


def _log_tender_error(title, tender_workspace_name, message=None):
	"""Log to Error Log, linked to the tender, and return the log's name."""
	log = frappe.log_error(
		title=title,
		message=_sanitize_log_text(message or frappe.get_traceback(), max_chars=4000),
		reference_doctype="Tender Workspace",
		reference_name=tender_workspace_name,
	)
	name = getattr(log, "name", None)
	_remember_tender_error(tender_workspace_name, name)
	return name


@frappe.whitelist(methods=["POST"])
def dismiss_background_job(tender_workspace_name, job_key):
	"""
	Remove a finished or failed entry from the Background Processes panel.

	Finished entries are kept for a day so they survive a reload, which means a
	stale failure can sit there contradicting the current state of the tender.
	Running jobs are refused so a live one cannot be hidden.
	"""
	frappe.has_permission("Tender Workspace", doc=tender_workspace_name, throw=True)

	jobs = _read_jobs(tender_workspace_name)
	job = jobs.get(job_key)
	if not job:
		return {"status": "Missing", "message": _("That process is no longer listed.")}
	if job.get("state") in ("queued", "running"):
		return {
			"status": "Running",
			"message": _("This process is still running and cannot be dismissed."),
		}

	jobs.pop(job_key, None)
	frappe.cache().set_value(
		_job_cache_key(tender_workspace_name), jobs, expires_in_sec=_JOB_CACHE_TTL
	)
	return {"status": "Dismissed", "message": _("Dismissed.")}


@frappe.whitelist()
def get_background_jobs(tender_workspace_name):
	"""
	Return the background jobs known for this tender, newest state first.

	Used on form load so the progress panel can be restored after a reload —
	realtime events alone are lost when the page is refreshed.
	"""
	if not tender_workspace_name:
		return {"jobs": []}
	frappe.has_permission("Tender Workspace", doc=tender_workspace_name, throw=True)

	from frappe.utils.background_jobs import is_job_enqueued

	jobs = _read_jobs(tender_workspace_name)
	changed = False
	for job_key, job in jobs.items():
		if job.get("state") not in ("queued", "running"):
			continue
		# Still queued or being worked on -> leave as is.
		if is_job_enqueued(_job_id(job_key, tender_workspace_name)):
			continue
		# Gone from RQ (worker crash/restart). Give it a grace period first —
		# `enqueue_after_commit` leaves a short window where the job is recorded
		# here but not yet in RQ. Reported only; never written back to the cache,
		# so a progress update racing with this read can't be clobbered.
		age = frappe.utils.time_diff_in_seconds(frappe.utils.now(), job.get("updated"))
		if age < _JOB_STALE_SECONDS:
			continue
		last_message = job.get("message")
		job["state"] = "failed"
		job["message"] = _("Stopped unexpectedly — check the Error Log and retry.")
		if not job.get("error_log"):
			job["error_log"] = _log_tender_error(
				"Tender background job stopped unexpectedly",
					tender_workspace_name,
					message=(
						f"Job key: {job_key}\n"
						f"Last message: {last_message}\n"
						f"Last update: {job.get('updated')}\n"
						"The RQ job is no longer enqueued and no completion event was recorded."
					),
			)
		changed = True

	if changed:
		frappe.cache().set_value(_job_cache_key(tender_workspace_name), jobs, expires_in_sec=_JOB_CACHE_TTL)

	return {"jobs": sorted(jobs.values(), key=lambda j: j.get("updated") or "", reverse=True)}


# Result statuses that mean "the operation did not produce a usable final result".
_FAILED_STATUSES = ("AI Failed", "AI Not Configured", "OCR Required", "No Items Found", "Failed")
_WARNING_STATUSES = ("Partial",)


def _job_state_for_result(result):
	status = (result or {}).get("status")
	if status in _WARNING_STATUSES:
		return "warning"
	if status in _FAILED_STATUSES:
		return "failed"
	return "done"


def _run_pipeline(tender_workspace_name, job_key, start_message, worker, *args):
	"""
	Run one AI step inside a background job and report its outcome to the form.

	`worker(tender_workspace_name, *args)` returns the same result dict the
	whitelisted endpoint used to return. A worker that already published its own
	closing event (the OCR pipelines do) says so with "final_published": True.
	"""
	try:
		_publish(tender_workspace_name, start_message, 10, job_key=job_key)
		result = worker(tender_workspace_name, *args) or {}
		if result.get("final_published"):
			return result
		state = _job_state_for_result(result)
		error_log = result.get("error_log")
		if state in ("failed", "warning") and not error_log:
			error_log = _latest_tender_error(tender_workspace_name)
		_publish(
			tender_workspace_name,
			result.get("message") or _("Completed."),
			100,
			reload=True,
			job_key=job_key,
			state=state,
			error_log=error_log,
		)
		return result
	except Exception:
		frappe.db.rollback()
		log = _log_tender_error(f"Tender background job failed: {job_key}", tender_workspace_name)
		_publish(
			tender_workspace_name,
			_("Failed — open the Error Log below, then retry."),
			100,
			reload=True,
			error_log=log,
			job_key=job_key,
			state="failed",
		)


def _enqueue_job(job_key, name, method, **kwargs):
	"""Queue a background job for a tender unless the same one is already running."""
	from frappe.utils.background_jobs import is_job_enqueued

	job_id = _job_id(job_key, name)
	if is_job_enqueued(job_id):
		return False

	_set_job_state(name, job_key, "queued", _("Queued…"), 0)
	frappe.enqueue(
		method,
		queue="long",
		timeout=1800,
		job_id=job_id,
		deduplicate=True,
		enqueue_after_commit=True,
		tender_workspace_name=name,
		**kwargs,
	)
	return True


# ---------------------------------------------------------------------------
# OCR text cache (file-based)
#   OCR is expensive (minutes), so the extracted text is cached. It is kept as
#   a private .txt File attached to the tender — NOT in a child-table field —
#   so the (large) text is never loaded on every `frappe.get_doc(...)` read.
#   The cache is loaded lazily, only inside the background pipeline.
# ---------------------------------------------------------------------------
def _ocr_cache_filename(td_name):
	return f"ocr-cache-{td_name}.txt"


def _read_ocr_cache(tender_name, td_name):
	"""Return cached OCR text for a tender document row, or '' if not cached."""
	names = frappe.get_all(
		"File",
		filters={
			"attached_to_doctype": "Tender Workspace",
			"attached_to_name": tender_name,
			"file_name": _ocr_cache_filename(td_name),
		},
		pluck="name",
	)
	if not names:
		return ""
	try:
		content = frappe.get_doc("File", names[0]).get_content()
		if isinstance(content, bytes):
			content = content.decode("utf-8", errors="ignore")
		return content or ""
	except Exception:
		_log_tender_error(
			"Tender OCR cache: read failed",
			tender_name,
			message=f"Row: {td_name}\nFile: {_ocr_cache_filename(td_name)}\n\n{frappe.get_traceback()}",
		)
		return ""


def _write_ocr_cache(tender_name, td_name, text):
	"""Cache OCR text as a private .txt file attached to the tender.

	Replaces any previous cache file for this document so re-runs don't pile up.
	"""
	filename = _ocr_cache_filename(td_name)
	existing = frappe.get_all(
		"File",
		filters={
			"attached_to_doctype": "Tender Workspace",
			"attached_to_name": tender_name,
			"file_name": filename,
		},
		pluck="name",
	)
	for name in existing:
		frappe.delete_doc("File", name, ignore_permissions=True, force=True)
	save_file(filename, (text or "").encode("utf-8"), "Tender Workspace", tender_name, is_private=1)


def _ocr_analyze_pipeline(tender_workspace_name):
	"""
	Background: OCR the tender document, then summarise it chunk-by-chunk.

	Concurrency note: the `doc` object is NOT held across the multi-minute AI
	loop. OCR text is cached to a private .txt File (not a child field); the doc
	is reloaded fresh immediately before the child-table mutation + save to avoid
	stale-save ("record changed") conflicts.
	"""
	doc = _get_doc(tender_workspace_name)
	tender_doc = _find_document(doc, TENDER_DOC_TYPES)
	if not tender_doc:
		return

	file_url = tender_doc.file
	source = tender_doc.file_name or file_url
	td_name = tender_doc.name

	try:
		_publish(doc.name, _("Running OCR on the document…"), 10)

		# 1) OCR (reuse the file cache). The text lives in a private .txt File,
		#    not on the doc, so it never bloats normal reads.
		ocr_text = _read_ocr_cache(doc.name, td_name)
		if len(ocr_text.strip()) < 100:
			ocr_text = document_parser.ocr_pdf_text(file_url)
			_write_ocr_cache(doc.name, td_name, ocr_text)
			frappe.db.commit()

		if len(ocr_text.strip()) < 100:
			frappe.db.set_value("Tender Document Item", td_name, {
				"ai_status": "OCR Required", "readable_status": "OCR Required",
				"ai_summary": _("OCR produced no readable text from this document."),
			}, update_modified=False)
			frappe.db.commit()
			error_log = _log_tender_error(
				"Tender OCR pipeline: empty OCR",
				doc.name,
				message=f"File: {source} ({file_url})",
			)
			_publish(doc.name, _("OCR produced no text."), 100, reload=True, state="failed", error_log=error_log)
			return

		# 2) Chunk + summarise (throttled). No DB writes inside the loop.
		chunks = _chunk_text(ocr_text)
		collected = []
		for idx, chunk in enumerate(chunks):
			rows = _ai_summary_rows(doc, chunk)
			if rows:
				collected.extend(rows)
			ai_service.throttle(ai_service.estimate_tokens(chunk))
			pct = 40 + int(50 * (idx + 1) / max(1, len(chunks)))
			_publish(doc.name, _("Analysed section {0}/{1}").format(idx + 1, len(chunks)), pct)

		if not collected:
			frappe.db.set_value("Tender Document Item", td_name, "ai_status", "Failed", update_modified=False)
			frappe.db.commit()
			error_log = _log_tender_error(
				"Tender OCR analyze pipeline: no summary rows",
				doc.name,
				message=f"File: {source} ({file_url})\nOCR text length: {len(ocr_text.strip())}\nChunk count: {len(chunks)}",
			)
			_publish(
				doc.name,
				_("OCR succeeded, but AI returned no summary rows."),
				100,
				reload=True,
				state="failed",
				error_log=error_log,
			)
			return

		# 3) Reload fresh, then persist rows (dedupe near-identical findings).
		doc = frappe.get_doc("Tender Workspace", tender_workspace_name)
		tender_doc = _find_document(doc, TENDER_DOC_TYPES)
		doc.ai_summary = [r for r in doc.ai_summary if r.source_document != source]
		seen = set()
		for row in collected:
			key = (row.get("summary_type"), (row.get("extracted_text") or "")[:60])
			if key in seen:
				continue
			seen.add(key)
			doc.append("ai_summary", {
				"summary_type": row.get("summary_type"),
				"extracted_text": row.get("extracted_text") or "",
				"source_document": source,
				"page_number": str(row.get("page_number") or ""),
				"confirmed": 0,
			})

		if tender_doc:
			tender_doc.ai_status = "Processed"
			tender_doc.readable_status = "Yes"
			tender_doc.ai_summary = _("AI analysis complete (OCR + AI).")
		doc.status = "AI Analyzed"
		_update_knowledge_cache_fields(doc, preferred_text_by_row={td_name: ocr_text})
		doc.save()
		frappe.db.commit()
		_publish(doc.name, _("Analysis complete."), 100, reload=True)
	except Exception:
		frappe.db.rollback()
		log = _log_tender_error("Tender OCR analyze pipeline failed", tender_workspace_name)
		frappe.db.set_value("Tender Document Item", td_name, "ai_status", "Failed", update_modified=False)
		frappe.db.commit()
		_publish(
			tender_workspace_name, _("Analysis failed."), 100,
			reload=True, state="failed", error_log=log,
		)


def _ocr_boq_pipeline(tender_workspace_name):
	"""Background: OCR the BOQ document and extract line items from the table pages."""
	doc = _get_doc(tender_workspace_name)
	boq_doc = _find_document(doc, BOQ_DOC_TYPES) or _find_document(doc, TENDER_DOC_TYPES)
	if not boq_doc:
		return

	file_url = boq_doc.file
	boq_name = boq_doc.name
	try:
		_publish(doc.name, _("Running OCR on the BOQ…"), 15, job_key="boq")

		# OCR per page so we can target only the BOQ / quantities pages.
		pages = document_parser.ocr_pdf_pages(file_url)
		if not any(t.strip() for _, t in pages):
			frappe.db.set_value("Tender Document Item", boq_name, {
				"ai_status": "OCR Required", "readable_status": "OCR Required",
			}, update_modified=False)
			frappe.db.commit()
			error_log = _log_tender_error(
				"Tender OCR BOQ: empty OCR",
				doc.name,
				message=f"File: {boq_doc.file_name or file_url} ({file_url})",
			)
			_publish(doc.name, _("OCR produced no text."), 100, reload=True, job_key="boq", state="failed", error_log=error_log)
			return

		boq_pages = [t for _, t in pages if any(m in t for m in _BOQ_MARKERS)]
		boq_text = "\n\n".join(boq_pages) if boq_pages else "\n\n".join(t for _, t in pages)

		_publish(doc.name, _("Extracting BOQ items…"), 60, job_key="boq")
		rows, _confidence, _issues = _extract_boq_from_text(boq_text[: _CHUNK_CHARS * 2])
		rows = _filter_boq_rows(rows)

		if not rows:
			frappe.db.set_value("Tender Document Item", boq_name, {
				"ai_status": "OCR Required", "readable_status": "OCR Required",
			}, update_modified=False)
			frappe.db.commit()
			error_log = _log_tender_error(
				"Tender OCR BOQ: no items extracted",
				doc.name,
				message=(
					f"File: {boq_doc.file_name or file_url} ({file_url})\n"
					f"OCR pages: {len(pages)}\nBOQ marker pages: {len(boq_pages)}\n"
					f"BOQ text length: {len(boq_text.strip())}"
				),
			)
			_publish(doc.name, _("Could not extract BOQ items."), 100, reload=True, job_key="boq", state="failed", error_log=error_log)
			return

		# Reload fresh before mutating the child table + saving.
		doc = frappe.get_doc("Tender Workspace", tender_workspace_name)
		boq_doc = _find_document(doc, BOQ_DOC_TYPES)
		doc.set("boq_items", [])
		for r in rows:
			doc.append("boq_items", r)
		if boq_doc:
			boq_doc.ai_status = "Extracted"
			boq_doc.readable_status = "Yes"
		doc.status = "BOQ Extracted"
		_update_knowledge_cache_fields(doc, preferred_text_by_row={boq_name: boq_text})
		doc.save()
		frappe.db.commit()
		_publish(doc.name, _("BOQ extracted ({0} items).").format(len(rows)), 100, reload=True, job_key="boq")
	except Exception:
		frappe.db.rollback()
		log = _log_tender_error("Tender OCR BOQ pipeline failed", tender_workspace_name)
		frappe.db.set_value("Tender Document Item", boq_name, "ai_status", "Failed", update_modified=False)
		frappe.db.commit()
		_publish(
			tender_workspace_name, _("BOQ extraction failed."), 100,
			reload=True, job_key="boq", state="failed", error_log=log,
		)
